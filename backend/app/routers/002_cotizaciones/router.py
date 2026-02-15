"""API routes for LATAM exchange rates (cotizaciones)."""
from datetime import date, datetime, timedelta
from typing import List, Dict, Optional
from io import BytesIO
from flask import Blueprint, request, jsonify, abort, send_file
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from ...database import execute_query, execute_query_single

bp = Blueprint('cotizaciones', __name__)


def parse_fecha(fecha_val):
    """Convierte fecha de SQLite (puede ser date o datetime string) a date."""
    if isinstance(fecha_val, date):
        return fecha_val
    fecha_str = str(fecha_val)
    # Si tiene hora (contiene espacio), tomar solo la parte de fecha
    if ' ' in fecha_str:
        fecha_str = fecha_str.split(' ')[0]
    return date.fromisoformat(fecha_str)


@bp.route('/cotizaciones', methods=['GET'])
def get_cotizaciones():
    """
    Obtiene cotizaciones diarias de tipos de cambio LATAM.
    
    Query params:
    - product_ids[]: Lista de IDs de productos (cotizaciones)
    - fecha_desde: Fecha inicial (YYYY-MM-DD)
    - fecha_hasta: Fecha final (YYYY-MM-DD)
    """
    try:
        # Obtener parámetros
        product_ids = request.args.getlist('product_ids[]', type=int)
        fecha_desde_str = request.args.get('fecha_desde')
        fecha_hasta_str = request.args.get('fecha_hasta')
        
        if not product_ids:
            return jsonify({'error': 'Se requiere al menos un product_id'}), 400
        
        if not fecha_desde_str or not fecha_hasta_str:
            return jsonify({'error': 'Se requieren fecha_desde y fecha_hasta'}), 400
        
        fecha_desde = date.fromisoformat(fecha_desde_str)
        fecha_hasta = date.fromisoformat(fecha_hasta_str)
        
        if fecha_desde > fecha_hasta:
            return jsonify({'error': 'fecha_desde debe ser anterior a fecha_hasta'}), 400
        
        result = []
        
        # Convertir product_ids sintéticos a (id_variable, id_pais) pairs
        fks_list = []
        for product_id in product_ids:
            id_variable = product_id // 10000
            id_pais = product_id % 10000
            fks_list.append((id_variable, id_pais))
        
        # Construir condiciones WHERE para (id_variable, id_pais) pairs
        fks_conditions = []
        fks_params = []
        for id_var, id_pais in fks_list:
            fks_conditions.append("(m.id_variable = ? AND m.id_pais = ?)")
            fks_params.extend([id_var, id_pais])
        
        # Obtener información de cada cotización
        try:
            query_products = f"""
                SELECT 
                    (m.id_variable * 10000 + m.id_pais) as id,
                    v.id_nombre_variable as nombre,
                    m.fuente,
                    v.id_variable,
                    pg.nombre_pais_grupo as pais
                FROM maestro m
                LEFT JOIN pais_grupo pg ON m.id_pais = pg.id_pais
                LEFT JOIN variables v ON m.id_variable = v.id_variable
                WHERE ({' OR '.join(fks_conditions)})
                AND (m.activo = 1 OR CAST(m.activo AS INTEGER) = 1)
            """
            products = execute_query(query_products, tuple(fks_params))
        except Exception as join_error:
            # Si el JOIN falla, intentar sin pais_grupo
            try:
                query_products = f"""
                    SELECT 
                        (m.id_variable * 10000 + m.id_pais) as id,
                        v.id_nombre_variable as nombre,
                        m.fuente,
                        v.id_variable,
                        NULL as pais
                    FROM maestro m
                    LEFT JOIN variables v ON m.id_variable = v.id_variable
                    WHERE ({' OR '.join(fks_conditions)})
                    AND (m.activo = 1 OR CAST(m.activo AS INTEGER) = 1)
                """
                products = execute_query(query_products, tuple(fks_params))
            except Exception:
                # Si variables no existe, usar solo maestro
                query_products = f"""
                    SELECT 
                        (id_variable * 10000 + id_pais) as id,
                        NULL as nombre,
                        fuente,
                        id_variable,
                        NULL as pais
                    FROM maestro
                    WHERE ({' OR '.join(fks_conditions)})
                    AND (activo = 1 OR CAST(activo AS INTEGER) = 1)
                """
                products = execute_query(query_products, tuple(fks_params))
        
        if not products:
            return jsonify({'error': 'No se encontraron cotizaciones activas'}), 404
        
        # Obtener datos de cada cotización
        for product in products:
            product_id = product['id']
            product_name = product['nombre']
            product_source = product.get('fuente', '')
            product_unidad = product.get('unidad', '')
            product_pais = product.get('pais', '')
            
            # Convertir id sintético a id_variable e id_pais
            id_variable = product_id // 10000
            id_pais = product_id % 10000
            
            # Obtener precios: ampliar rango hacia atrás 250 días para calcular variaciones
            fecha_desde_eff = min(fecha_desde, fecha_hasta - timedelta(days=250))
            fecha_desde_str = fecha_desde.isoformat()
            fecha_hasta_str = fecha_hasta.isoformat()
            fecha_desde_eff_str = fecha_desde_eff.isoformat()
            query_prices = """
                SELECT fecha, valor
                FROM maestro_precios
                WHERE id_variable = ? AND id_pais = ? 
                AND DATE(fecha) >= DATE(?)
                AND DATE(fecha) <= DATE(?)
                ORDER BY fecha ASC
            """
            prices = execute_query(query_prices, (id_variable, id_pais, fecha_desde_eff_str, fecha_hasta_str))
            
            if not prices:
                continue
            
            # Normalizar fechas y preparar datos (para el gráfico solo el rango pedido)
            data_all = []
            for price_item in prices:
                fecha_obj = parse_fecha(price_item['fecha'])
                data_all.append({
                    'fecha': fecha_obj.isoformat(),
                    'valor': float(price_item['valor'])
                })
            # Datos para el gráfico: solo [fecha_desde, fecha_hasta]
            data = [d for d in data_all if fecha_desde <= parse_fecha(d['fecha']) <= fecha_hasta]
            
            # Calcular resumen (intervalo pedido)
            if data:
                precio_inicial = data[0]['valor']
                precio_final = data[-1]['valor']
                variacion = 0.0
                if precio_inicial > 0:
                    variacion = ((precio_final - precio_inicial) / precio_inicial) * 100
                fecha_inicial = data[0]['fecha']
                fecha_final = data[-1]['fecha']
            else:
                precio_inicial = None
                precio_final = None
                variacion = 0.0
                fecha_inicial = None
                fecha_final = None
            
            # Variaciones 1d, 5d (1 sem), 22d (1 mes), 250d (1 año): N observaciones atrás (días con dato), no N días naturales
            fecha_max = fecha_final  # última fecha en el rango
            variacion_1d = variacion_5d = variacion_22d = variacion_250d = None
            if data_all and fecha_max:
                # Lista ordenada por fecha para tomar "N observaciones atrás"
                sorted_all = sorted(data_all, key=lambda d: d['fecha'])
                fechas_ordenadas = [d['fecha'] for d in sorted_all]
                try:
                    idx_max = fechas_ordenadas.index(fecha_max)
                except ValueError:
                    idx_max = len(fechas_ordenadas) - 1
                v0 = sorted_all[idx_max]['valor'] if idx_max >= 0 else None
                if v0 is not None and v0 > 0:
                    for n_obs, key in [(1, 'variacion_1d'), (5, 'variacion_5d'), (22, 'variacion_22d'), (250, 'variacion_250d')]:
                        idx_ant = idx_max - n_obs
                        if idx_ant >= 0:
                            v_ant = sorted_all[idx_ant]['valor']
                            if v_ant is not None and v_ant > 0:
                                val = ((v0 - v_ant) / v_ant) * 100
                                if key == 'variacion_1d': variacion_1d = round(val, 2)
                                elif key == 'variacion_5d': variacion_5d = round(val, 2)
                                elif key == 'variacion_22d': variacion_22d = round(val, 2)
                                else: variacion_250d = round(val, 2)
            
            # Obtener id_variable del producto
            product_id_variable = product.get('id_variable')
            
            result.append({
                'product_id': product_id,
                'product_name': product_name,
                'product_source': product_source,
                'unidad': product_unidad,
                'pais': product_pais,
                'id_variable': product_id_variable if product_id_variable is not None else None,
                'data': data,
                'summary': {
                    'precio_inicial': precio_inicial,
                    'precio_final': precio_final,
                    'variacion': variacion,
                    'fecha_inicial': fecha_inicial,
                    'fecha_final': fecha_final,
                    'fecha_max': fecha_max,
                    'variacion_1d': variacion_1d,
                    'variacion_5d': variacion_5d,
                    'variacion_22d': variacion_22d,
                    'variacion_250d': variacion_250d
                }
            })
        
        return jsonify(result)
    
    except ValueError as e:
        return jsonify({'error': f'Error en formato de fecha: {str(e)}'}), 400
    except Exception as e:
        return jsonify({'error': f'Error al obtener cotizaciones: {str(e)}'}), 500


@bp.route('/cotizaciones/products', methods=['GET'])
def get_cotizaciones_products():
    """
    Obtiene la lista de cotizaciones disponibles.
    Filtra por países configurados en filtros_graph_pais para id_graph=2 (Cotizaciones).
    Solo incluye id_variable 20 (oficial), 21 (no oficial) y 85 (sintético).
    """
    try:
        # Obtener países permitidos desde filtros_graph_pais para graph id=2 (Cotizaciones)
        try:
            # Filtrar por países configurados en filtros_graph_pais
            try:
                # JOIN con variables, pais_grupo y filtro por países permitidos usando subquery
                query = """
                    SELECT DISTINCT
                        (m.id_variable * 10000 + m.id_pais) as id,
                        v.id_nombre_variable as nombre,
                        m.fuente,
                        m.periodicidad,
                        pg.nombre_pais_grupo as pais,
                        v.id_variable
                    FROM maestro m
                    LEFT JOIN pais_grupo pg ON m.id_pais = pg.id_pais
                    LEFT JOIN variables v ON m.id_variable = v.id_variable
                    WHERE m.periodicidad = 'D'
                    AND m.id_variable IN (20, 21, 85)
                    AND (m.activo = 1 OR CAST(m.activo AS INTEGER) = 1)
                    AND m.id_pais IN (
                        SELECT DISTINCT f.id_pais
                        FROM filtros_graph_pais f
                        WHERE f.id_graph = 2
                    )
                    ORDER BY pg.nombre_pais_grupo, v.id_nombre_variable
                """
                results = execute_query(query)
            except Exception as join_error:
                # Si el JOIN o la subquery falla, intentar solo con variables
                try:
                    query = """
                        SELECT DISTINCT
                            (m.id_variable * 10000 + m.id_pais) as id,
                            v.id_nombre_variable as nombre,
                            m.fuente,
                            m.periodicidad,
                            NULL as pais,
                            v.id_variable
                        FROM maestro m
                        LEFT JOIN variables v ON m.id_variable = v.id_variable
                        WHERE m.periodicidad = 'D'
                        AND m.id_variable IN (20, 21, 85)
                        AND (m.activo = 1 OR CAST(m.activo AS INTEGER) = 1)
                        AND m.id_pais IN (
                            SELECT DISTINCT f.id_pais
                            FROM filtros_graph_pais f
                            WHERE f.id_graph = 2
                        )
                        ORDER BY v.id_nombre_variable
                    """
                    results = execute_query(query)
                except Exception:
                    # Si filtros_graph_pais no existe o falla, devolver todos sin filtro
                    try:
                        query = """
                            SELECT DISTINCT
                                (m.id_variable * 10000 + m.id_pais) as id,
                                v.id_nombre_variable as nombre,
                                m.fuente,
                                m.periodicidad,
                                pg.nombre_pais_grupo as pais,
                                v.id_variable
                            FROM maestro m
                            LEFT JOIN pais_grupo pg ON m.id_pais = pg.id_pais
                            LEFT JOIN variables v ON m.id_variable = v.id_variable
                            WHERE m.periodicidad = 'D'
                            AND m.id_variable IN (20, 21, 85)
                            AND (m.activo = 1 OR CAST(m.activo AS INTEGER) = 1)
                            ORDER BY pais, v.id_nombre_variable
                        """
                        results = execute_query(query)
                    except Exception:
                        query = """
                            SELECT DISTINCT
                                (id_variable * 10000 + id_pais) as id,
                                NULL as nombre,
                                fuente,
                                periodicidad,
                                NULL as pais,
                                id_variable
                            FROM maestro
                            WHERE periodicidad = 'D'
                            AND id_variable IN (20, 21, 85)
                            AND (activo = 1 OR CAST(activo AS INTEGER) = 1)
                            ORDER BY id_variable
                        """
                        results = execute_query(query)
        except Exception:
            # Si todo falla, usar lógica sin filtro
            try:
                query = """
                    SELECT DISTINCT
                        (m.id_variable * 10000 + m.id_pais) as id,
                        v.id_nombre_variable as nombre,
                        m.fuente,
                        m.periodicidad,
                        pg.nombre_pais_grupo as pais,
                        v.id_variable
                    FROM maestro m
                    LEFT JOIN pais_grupo pg ON m.id_pais = pg.id_pais
                    LEFT JOIN variables v ON m.id_variable = v.id_variable
                    WHERE m.periodicidad = 'D'
                    AND m.id_variable IN (20, 21, 85)
                    AND (m.activo = 1 OR CAST(m.activo AS INTEGER) = 1)
                    ORDER BY pais, v.id_nombre_variable
                """
                results = execute_query(query)
            except Exception:
                query = """
                    SELECT DISTINCT
                        (id_variable * 10000 + id_pais) as id,
                        NULL as nombre,
                        fuente,
                        periodicidad,
                        NULL as pais,
                        id_variable
                    FROM maestro
                    WHERE periodicidad = 'D'
                    AND id_variable IN (20, 21, 85)
                    AND (activo = 1 OR CAST(activo AS INTEGER) = 1)
                    ORDER BY id_variable
                """
                results = execute_query(query)
        return jsonify(results)
    
    except Exception as e:
        return jsonify({'error': f'Error al obtener productos: {str(e)}'}), 500


@bp.route('/cotizaciones/export', methods=['GET'])
def export_cotizaciones_to_excel():
    """
    Exporta cotizaciones a Excel.
    
    Query params:
    - product_ids[]: Lista de IDs de productos
    - fecha_desde: Fecha inicial (YYYY-MM-DD)
    - fecha_hasta: Fecha final (YYYY-MM-DD)
    """
    try:
        product_ids = request.args.getlist('product_ids[]', type=int)
        fecha_desde_str = request.args.get('fecha_desde')
        fecha_hasta_str = request.args.get('fecha_hasta')
        
        if not product_ids or not fecha_desde_str or not fecha_hasta_str:
            return jsonify({'error': 'Parámetros incompletos'}), 400
        
        fecha_desde = date.fromisoformat(fecha_desde_str)
        fecha_hasta = date.fromisoformat(fecha_hasta_str)
        
        # Crear workbook
        wb = Workbook()
        wb.remove(wb.active)  # Remover hoja por defecto
        
        # Hoja 1: Cotizaciones
        ws_cotizaciones = wb.create_sheet("Cotizaciones")
        
        # Encabezados
        headers = ['Fecha', 'País/Cotización', 'Valor', 'Unidad', 'Fuente']
        ws_cotizaciones.append(headers)
        
        # Estilo para encabezados
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for col in range(1, len(headers) + 1):
            cell = ws_cotizaciones.cell(row=1, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Convertir product_ids sintéticos a (id_variable, id_pais) pairs
        fks_list = []
        for product_id in product_ids:
            id_variable = product_id // 10000
            id_pais = product_id % 10000
            fks_list.append((id_variable, id_pais))
        
        # Construir condiciones WHERE para (id_variable, id_pais) pairs
        fks_conditions = []
        fks_params = []
        for id_var, id_pais in fks_list:
            fks_conditions.append("(m.id_variable = ? AND m.id_pais = ?)")
            fks_params.extend([id_var, id_pais])
        
        # Obtener datos
        query_products = f"""
            SELECT 
                (m.id_variable * 10000 + m.id_pais) as id,
                v.id_nombre_variable as nombre,
                m.fuente
            FROM maestro m
            LEFT JOIN variables v ON m.id_variable = v.id_variable
            WHERE ({' OR '.join(fks_conditions)})
            AND m.activo = 1
        """
        products = execute_query(query_products, tuple(fks_params))
        
        row_num = 2
        for product in products:
            product_id = product['id']
            product_name = product['nombre']
            product_source = product.get('fuente', '')
            
            # Convertir id sintético a id_variable e id_pais
            id_variable = product_id // 10000
            id_pais = product_id % 10000
            
            query_prices = """
                SELECT fecha, valor
                FROM maestro_precios
                WHERE id_variable = ? AND id_pais = ? AND fecha >= ? AND fecha <= ?
                ORDER BY fecha ASC
            """
            prices = execute_query(query_prices, (id_variable, id_pais, fecha_desde, fecha_hasta))
            
            for price_item in prices:
                fecha_obj = parse_fecha(price_item['fecha'])
                ws_cotizaciones.append([
                    fecha_obj.isoformat(),
                    product_name,
                    float(price_item['valor']),
                    '',  # unidad no existe en nuevo schema
                    product_source
                ])
                row_num += 1
        
        # Ajustar ancho de columnas
        ws_cotizaciones.column_dimensions['A'].width = 12
        ws_cotizaciones.column_dimensions['B'].width = 30
        ws_cotizaciones.column_dimensions['C'].width = 15
        ws_cotizaciones.column_dimensions['D'].width = 15
        ws_cotizaciones.column_dimensions['E'].width = 20
        
        # Hoja 2: Metadatos
        ws_metadata = wb.create_sheet("Metadatos")
        ws_metadata.append(['Campo', 'Valor'])
        
        # Estilo para encabezados de metadatos
        for col in range(1, 3):
            cell = ws_metadata.cell(row=1, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        ws_metadata.append(['Fecha desde', fecha_desde_str])
        ws_metadata.append(['Fecha hasta', fecha_hasta_str])
        ws_metadata.append(['Total cotizaciones', len(products)])
        ws_metadata.append(['', ''])
        ws_metadata.append(['Cotización', 'ID'])
        
        for product in products:
            ws_metadata.append([product['nombre'], product['id']])
        
        ws_metadata.column_dimensions['A'].width = 25
        ws_metadata.column_dimensions['B'].width = 15
        
        # Guardar en BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Generar nombre de archivo
        filename = f"cotizaciones_latam_{fecha_desde_str}_{fecha_hasta_str}.xlsx"
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
    
    except ValueError as e:
        return jsonify({'error': f'Error en formato de fecha: {str(e)}'}), 400
    except Exception as e:
        return jsonify({'error': f'Error al exportar: {str(e)}'}), 500
