"""API routes for Dollar Inflation (Inflación en dólares)."""
import importlib
from datetime import date, datetime
from typing import List, Dict, Optional
from flask import Blueprint, request, jsonify, send_file
from ...database import execute_query, execute_query_single
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
import io

# Import from numbered module using importlib
_dcp_module = importlib.import_module('app.routers.001_dcp.router')
convert_to_monthly = _dcp_module.convert_to_monthly

bp = Blueprint('inflacion_dolares', __name__)


def get_ipc_by_country(id_pais: int, fecha_desde: date, fecha_hasta: date) -> Dict[date, float]:
    """
    Obtiene el IPC mensual de un país específico.
    
    Args:
        id_pais: ID del país
        fecha_desde: Fecha inicial
        fecha_hasta: Fecha final
    
    Returns:
        Dict con fechas (primer día del mes) como keys y valores IPC como values
    """
    # Buscar IPC del país: id_variable = 9 (IPC), periodicidad='M', id_pais
    query_maestro = """
        SELECT m.id_variable, m.id_pais, m.periodicidad
        FROM maestro m
        WHERE m.id_pais = ?
        AND m.id_variable = 9
        AND m.periodicidad = 'M'
        LIMIT 1
    """
    
    # Buscar registro de maestro
    maestro_info = execute_query_single(query_maestro, (id_pais,))
    
    if not maestro_info:
        print(f"[DEBUG] get_ipc_by_country: No se encontró registro en maestro para id_pais={id_pais}, id_variable=9, periodicidad='M'")
        return {}
    
    id_variable = maestro_info['id_variable']
    id_pais_maestro = maestro_info['id_pais']
    periodicidad = maestro_info['periodicidad']
    
    print(f"[DEBUG] get_ipc_by_country: Encontrado maestro - id_variable={id_variable}, id_pais={id_pais_maestro}, periodicidad={periodicidad}")
    
    # Obtener datos del IPC
    query = """
        SELECT fecha, valor
        FROM maestro_precios
        WHERE id_variable = ? AND id_pais = ? 
        AND DATE(fecha) >= DATE(?)
        AND DATE(fecha) <= DATE(?)
        ORDER BY fecha ASC
    """
    
    fecha_desde_str = fecha_desde.isoformat()
    fecha_hasta_str = fecha_hasta.isoformat()
    
    print(f"[DEBUG] get_ipc_by_country: Buscando datos para id_pais={id_pais}, id_variable={id_variable}, rango {fecha_desde_str} a {fecha_hasta_str}")
    
    try:
        raw_data = execute_query(query, (id_variable, id_pais_maestro, fecha_desde_str, fecha_hasta_str))
        print(f"[DEBUG] get_ipc_by_country: Query ejecutada - encontrados {len(raw_data) if raw_data else 0} registros en rango {fecha_desde_str} a {fecha_hasta_str}")
        if raw_data:
            fechas_encontradas = sorted(set([item['fecha'] for item in raw_data]))
            print(f"[DEBUG] get_ipc_by_country: Fechas encontradas en raw_data: {fechas_encontradas[0] if fechas_encontradas else 'N/A'} a {fechas_encontradas[-1] if fechas_encontradas else 'N/A'}")
    except Exception as e:
        print(f"[ERROR] get_ipc_by_country: Error al ejecutar query: {str(e)}")
        return {}
    
    if not raw_data:
        print(f"[DEBUG] get_ipc_by_country: No hay datos en maestro_precios para id_variable={id_variable}, id_pais={id_pais_maestro} en el rango especificado")
        return {}
    
    # Convertir a mensual (ya es mensual, pero puede haber múltiples valores por mes)
    monthly_data = convert_to_monthly(raw_data, periodicidad)
    
    # Filtrar por rango usando comparación de año-mes
    fecha_desde_ym = (fecha_desde.year, fecha_desde.month)
    fecha_hasta_ym = (fecha_hasta.year, fecha_hasta.month)
    
    filtered_data = {}
    for item in monthly_data:
        mes_fecha = item['fecha']
        if not isinstance(mes_fecha, date):
            if isinstance(mes_fecha, str):
                mes_fecha = date.fromisoformat(mes_fecha)
            else:
                continue
        
        mes_fecha_ym = (mes_fecha.year, mes_fecha.month)
        if mes_fecha_ym >= fecha_desde_ym and mes_fecha_ym <= fecha_hasta_ym:
            filtered_data[mes_fecha] = item['valor']
    
    return filtered_data


def get_all_tc_monthly(id_paises: List[int], fecha_desde: date, fecha_hasta: date) -> Dict[int, Dict[date, float]]:
    """
    Obtiene los tipos de cambio mensuales para los países especificados.
    Calcula el promedio mensual para cada país.
    Más eficiente que llamar get_tc_by_country() múltiples veces.
    
    Args:
        id_paises: Lista de id_pais de los países seleccionados
        fecha_desde: Fecha inicial
        fecha_hasta: Fecha final
    
    Returns:
        Dict con id_pais como key y Dict[date, float] como value (TC mensual por país)
    """
    if not id_paises:
        return {}
    
    # Una sola query para obtener todos los TC (id_variable=20) de los países seleccionados
    placeholders = ','.join(['?'] * len(id_paises))
    query = f"""
        SELECT 
            mp.id_pais,
            mp.fecha,
            mp.valor
        FROM maestro_precios mp
        INNER JOIN maestro m ON mp.id_variable = m.id_variable AND mp.id_pais = m.id_pais
        WHERE mp.id_variable = 20
        AND mp.id_pais IN ({placeholders})
        AND DATE(mp.fecha) >= DATE(?)
        AND DATE(mp.fecha) <= DATE(?)
        ORDER BY mp.id_pais, mp.fecha ASC
    """
    
    fecha_desde_str = fecha_desde.isoformat()
    fecha_hasta_str = fecha_hasta.isoformat()
    
    try:
        raw_data = execute_query(query, tuple(id_paises + [fecha_desde_str, fecha_hasta_str]))
    except Exception as e:
        print(f"[ERROR] get_all_tc_monthly: Error al obtener datos: {str(e)}")
        return {}
    
    if not raw_data:
        return {}
    
    # Agrupar por país
    tc_by_pais = {}
    for row in raw_data:
        id_pais = row['id_pais']
        if id_pais not in tc_by_pais:
            tc_by_pais[id_pais] = []
        tc_by_pais[id_pais].append({
            'fecha': row['fecha'],
            'valor': row['valor']
        })
    
    # Convertir a mensual para cada país y filtrar por rango
    resultado = {}
    fecha_desde_ym = (fecha_desde.year, fecha_desde.month)
    fecha_hasta_ym = (fecha_hasta.year, fecha_hasta.month)
    
    # id_variable=20 siempre es diaria (periodicidad='D')
    periodicidad = 'D'
    
    for id_pais, datos in tc_by_pais.items():
        # Convertir a mensual
        monthly_data = convert_to_monthly(datos, periodicidad)
        
        # Filtrar por rango
        filtered_data = {}
        for item in monthly_data:
            mes_fecha = item['fecha']
            if not isinstance(mes_fecha, date):
                if isinstance(mes_fecha, str):
                    mes_fecha = date.fromisoformat(mes_fecha)
                else:
                    continue
            
            mes_fecha_ym = (mes_fecha.year, mes_fecha.month)
            if mes_fecha_ym >= fecha_desde_ym and mes_fecha_ym <= fecha_hasta_ym:
                filtered_data[mes_fecha] = item['valor']
        
        resultado[id_pais] = filtered_data
    
    return resultado


@bp.route('/inflacion-dolares/products', methods=['GET'])
def get_inflacion_dolares_products():
    """
    Obtiene la lista de países disponibles con cotización e IPC.
    Filtra por países configurados en filtros_graph_pais para id_graph=3 (Inflación en dólares).
    """
    try:
        # Primero obtener los países permitidos desde filtros_graph_pais
        try:
            paises_permitidos_query = """
                SELECT DISTINCT f.id_pais, pg.nombre_pais_grupo
                FROM filtros_graph_pais f
                LEFT JOIN pais_grupo pg ON f.id_pais = pg.id_pais
                WHERE f.id_graph = 3
                ORDER BY pg.nombre_pais_grupo
            """
            paises_permitidos = execute_query(paises_permitidos_query)
        except Exception as e:
            print(f"[ERROR] inflacion-dolares/products: Error al obtener países permitidos: {str(e)}")
            return jsonify({'error': 'No se pudo obtener países permitidos desde filtros_graph_pais'}), 500
        
        if not paises_permitidos:
            print("[DEBUG] inflacion-dolares/products: No hay países configurados en filtros_graph_pais para graph 3")
            return jsonify([])
        
        print(f"[DEBUG] inflacion-dolares/products: {len(paises_permitidos)} países en filtros_graph_pais para graph 3")
        
        results = []
        paises_procesados = set()  # Para evitar duplicados por id_pais
        
        # Para cada país permitido, buscar una cotización y verificar IPC
        for pais_filtro in paises_permitidos:
            id_pais = pais_filtro['id_pais']
            nombre_pais = pais_filtro.get('nombre_pais_grupo', '')
            
            if not nombre_pais or id_pais in paises_procesados:
                continue
            
            # Buscar cotización USD/LC para este país (id_variable = 20, siempre diaria)
            query_cotizacion = """
                SELECT 
                    (m.id_variable * 10000 + m.id_pais) as id,
                    v.id_nombre_variable as nombre,
                    pg.nombre_pais_grupo as pais,
                    m.fuente
                FROM maestro m
                LEFT JOIN variables v ON m.id_variable = v.id_variable
                LEFT JOIN pais_grupo pg ON m.id_pais = pg.id_pais
                WHERE m.id_pais = ?
                AND m.id_variable = 20
                AND (m.activo = 1 OR CAST(m.activo AS INTEGER) = 1)
                LIMIT 1
            """
            cotizacion = execute_query_single(query_cotizacion, (id_pais,))
            
            if not cotizacion:
                print(f"[DEBUG] inflacion-dolares/products: No se encontró cotización para país id={id_pais} ({nombre_pais})")
                continue
            
            # Verificar si existe IPC para este país
            # Buscar por id_variable = 9 (IPC) y id_pais, y verificar que tenga datos
            query_ipc = """
                SELECT COUNT(mp.fecha) as cantidad_datos
                FROM maestro ipc
                LEFT JOIN maestro_precios mp ON ipc.id_variable = mp.id_variable AND ipc.id_pais = mp.id_pais
                WHERE ipc.id_pais = ?
                AND ipc.id_variable = 9
                AND ipc.periodicidad = 'M'
            """
            ipc_result = execute_query_single(query_ipc, (id_pais,))
            
            cantidad_datos = ipc_result.get('cantidad_datos', 0) if ipc_result else 0
            
            if cantidad_datos > 0:
                results.append(cotizacion)
                paises_procesados.add(id_pais)
                print(f"[DEBUG] inflacion-dolares/products: Agregado país id={id_pais} ({nombre_pais}) - IPC con {cantidad_datos} datos")
            else:
                print(f"[DEBUG] inflacion-dolares/products: No se encontró IPC con datos para país id={id_pais} ({nombre_pais})")
        
        print(f"[DEBUG] inflacion-dolares/products: {len(results)} países con cotización e IPC disponible")
        return jsonify(results)
    
    except Exception as e:
        import traceback
        error_trace = traceback.format_exc()
        print(f"[ERROR] inflacion-dolares/products: {str(e)}")
        print(f"[TRACEBACK] {error_trace}")
        return jsonify({'error': f'Error al obtener productos: {str(e)}'}), 500


@bp.route('/inflacion-dolares', methods=['GET'])
def get_inflacion_dolares():
    """
    Calcula la inflación en dólares para países seleccionados.
    
    Fórmula: (IPC mensual) / (TC USD/LC mensual)
    Normalizado a base 100.
    
    Query params:
    - product_ids[]: Lista de IDs de cotizaciones (países)
    - fecha_desde: Fecha inicial (YYYY-MM-DD)
    - fecha_hasta: Fecha final (YYYY-MM-DD)
    """
    try:
        product_ids = request.args.getlist('product_ids[]', type=int)
        fecha_desde_str = request.args.get('fecha_desde')
        fecha_hasta_str = request.args.get('fecha_hasta')
        
        if not product_ids:
            return jsonify({'error': 'Se requiere al menos un product_id'}), 400
        
        if not fecha_desde_str or not fecha_hasta_str:
            return jsonify({'error': 'Se requieren fecha_desde y fecha_hasta'}), 400
        
        fecha_desde = date.fromisoformat(fecha_desde_str)
        fecha_hasta = date.fromisoformat(fecha_hasta_str)
        
        print(f"[DEBUG] inflacion-dolares: Request recibido - fecha_desde={fecha_desde}, fecha_hasta={fecha_hasta}")
        
        if fecha_desde > fecha_hasta:
            return jsonify({'error': 'fecha_desde debe ser anterior a fecha_hasta'}), 400
        
        result = []
        
        # Convertir product_ids sintéticos a (id_variable, id_pais) pairs
        fks_list = []
        for product_id in product_ids:
            id_variable = product_id // 10000
            id_pais = product_id % 10000
            fks_list.append((id_variable, id_pais))
        
        # Construir condiciones WHERE para (id_variable, id_pais) pairs
        fks_conditions = []
        fks_params = []
        for id_var, id_pais in fks_list:
            fks_conditions.append("(m.id_variable = ? AND m.id_pais = ?)")
            fks_params.extend([id_var, id_pais])
        
        # Obtener información de cada cotización/país
        query_products = f"""
            SELECT 
                (m.id_variable * 10000 + m.id_pais) as id,
                v.id_nombre_variable as nombre,
                pg.nombre_pais_grupo as pais,
                m.fuente
            FROM maestro m
            LEFT JOIN variables v ON m.id_variable = v.id_variable
            LEFT JOIN pais_grupo pg ON m.id_pais = pg.id_pais
            WHERE ({' OR '.join(fks_conditions)})
            AND (m.activo = 1 OR CAST(m.activo AS INTEGER) = 1)
        """
        products = execute_query(query_products, tuple(fks_params))
        
        if not products:
            return jsonify({'error': 'No se encontraron cotizaciones activas'}), 404
        
        # Extraer id_pais de los países seleccionados
        id_paises_seleccionados = [id_pais for _, id_pais in fks_list]
        
        # Obtener todos los TC de una vez (más eficiente: 1 query en lugar de N)
        all_tc_monthly = get_all_tc_monthly(id_paises_seleccionados, fecha_desde, fecha_hasta)
        print(f"[DEBUG] inflacion-dolares: TC obtenidos para {len(all_tc_monthly)} países")
        
        # Procesar cada país
        for product in products:
            product_id = product['id']
            product_name = product['nombre']
            pais = product.get('pais', '')
            product_source = product.get('fuente', '')
            product_unidad = product.get('unidad', '')
            
            # Obtener TC mensual del país desde el dict
            id_pais_product = product_id % 10000
            tc_monthly = all_tc_monthly.get(id_pais_product, {})
            
            if not tc_monthly:
                print(f"[DEBUG] inflacion-dolares: No se encontró TC para id_pais={id_pais_product} ({pais})")
                continue
            
            print(f"[DEBUG] inflacion-dolares: TC encontrado para {pais}: {len(tc_monthly)} meses")
            
            # Obtener IPC mensual del país
            ipc_monthly = get_ipc_by_country(id_pais_product, fecha_desde, fecha_hasta)
            
            if not ipc_monthly:
                print(f"[DEBUG] inflacion-dolares: No se encontró IPC para id_pais={id_pais_product} ({pais})")
                continue
            
            print(f"[DEBUG] inflacion-dolares: IPC encontrado para {pais}: {len(ipc_monthly)} meses")
            
            # Debug: mostrar fechas disponibles
            if tc_monthly:
                tc_fechas = sorted(tc_monthly.keys())
                print(f"[DEBUG] inflacion-dolares: TC fechas disponibles: {tc_fechas[0] if tc_fechas else 'N/A'} a {tc_fechas[-1] if tc_fechas else 'N/A'}")
            if ipc_monthly:
                ipc_fechas = sorted(ipc_monthly.keys())
                print(f"[DEBUG] inflacion-dolares: IPC fechas disponibles: {ipc_fechas[0] if ipc_fechas else 'N/A'} a {ipc_fechas[-1] if ipc_fechas else 'N/A'}")
            
            # Calcular índice: (IPC / TC)
            indices = []
            for mes_fecha, tc_valor in tc_monthly.items():
                if mes_fecha in ipc_monthly:
                    ipc_valor = ipc_monthly[mes_fecha]
                    if tc_valor > 0:  # Evitar división por cero en TC
                        indice_valor = ipc_valor / tc_valor
                        indices.append({'fecha': mes_fecha, 'valor': indice_valor})
                else:
                    print(f"[DEBUG] inflacion-dolares: Mes {mes_fecha} tiene TC pero no IPC")
            
            # Verificar meses con IPC pero sin TC
            for mes_fecha in ipc_monthly.keys():
                if mes_fecha not in tc_monthly:
                    print(f"[DEBUG] inflacion-dolares: Mes {mes_fecha} tiene IPC pero no TC")
            
            if len(indices) < 2:
                continue
            
            # Ordenar por fecha
            indices.sort(key=lambda x: x['fecha'])
            
            # Normalizar a base 100 (primer valor = 100)
            first_value = indices[0]['valor']
            if first_value == 0 or first_value is None:
                continue
            
            factor = 100.0 / first_value
            indices_normalized = [
                {
                    'fecha': idx['fecha'].isoformat() if isinstance(idx['fecha'], date) else idx['fecha'],
                    'valor': idx['valor'] * factor
                }
                for idx in indices
            ]
            
            # Calcular variaciones
            indice_inicial = indices_normalized[0]['valor']
            indice_final = indices_normalized[-1]['valor']
            variacion_indice = ((indice_final / indice_inicial) - 1.0) * 100 if indice_inicial > 0 else 0.0
            
            # Variación TC (usar fechas originales de indices, no normalizados)
            fecha_inicial_obj = indices[0]['fecha']
            fecha_final_obj = indices[-1]['fecha']
            tc_inicial = tc_monthly[fecha_inicial_obj]
            tc_final = tc_monthly[fecha_final_obj]
            variacion_tc = ((tc_final / tc_inicial) - 1.0) * 100 if tc_inicial > 0 else 0.0
            
            # Variación IPC (inflación)
            ipc_inicial = ipc_monthly[fecha_inicial_obj]
            ipc_final = ipc_monthly[fecha_final_obj]
            variacion_ipc = ((ipc_final / ipc_inicial) - 1.0) * 100 if ipc_inicial > 0 else 0.0
            
            # Fechas del intervalo real (convertir a ISO string)
            fecha_inicial = fecha_inicial_obj.isoformat() if isinstance(fecha_inicial_obj, date) else fecha_inicial_obj
            fecha_final = fecha_final_obj.isoformat() if isinstance(fecha_final_obj, date) else fecha_final_obj
            
            result.append({
                'product_id': product_id,
                'product_name': product_name,
                'pais': pais,
                'product_source': product_source,
                'unidad': product_unidad,
                'data': indices_normalized,
                'summary': {
                    'indice_inicial': indice_inicial,
                    'indice_final': indice_final,
                    'variacion_indice': variacion_indice,  # Inflación en dólares
                    'variacion_tc': variacion_tc,
                    'variacion_ipc': variacion_ipc,
                    'fecha_inicial': fecha_inicial,
                    'fecha_final': fecha_final
                }
            })
        
        return jsonify(result)
    
    except ValueError as e:
        return jsonify({'error': f'Error en formato de fecha: {str(e)}'}), 400
    except Exception as e:
        return jsonify({'error': f'Error al calcular inflación en dólares: {str(e)}'}), 500


@bp.route('/inflacion-dolares/export', methods=['GET'])
def export_inflacion_dolares_to_excel():
    """
    Exporta inflación en dólares a Excel con 3 hojas.
    
    Parámetros: iguales a /inflacion-dolares
    Retorna: archivo Excel
    
    Hojas:
    1. Índice Normalizado: Inflación en dólares normalizada a base 100
    2. Índice Original: IPC/TC sin normalizar
    3. Datos Originales: IPC y TC mensuales tal cual en la base de datos
    """
    try:
        product_ids = request.args.getlist('product_ids[]', type=int)
        
        if not product_ids:
            return jsonify({'error': 'Se requiere al menos un product_id'}), 400
        
        fecha_desde_str = request.args.get('fecha_desde')
        fecha_hasta_str = request.args.get('fecha_hasta')
        
        if not fecha_desde_str or not fecha_hasta_str:
            return jsonify({'error': 'Se requieren fecha_desde y fecha_hasta'}), 400
        
        fecha_desde = date.fromisoformat(fecha_desde_str)
        fecha_hasta = date.fromisoformat(fecha_hasta_str)
        
        if fecha_desde > fecha_hasta:
            return jsonify({'error': 'fecha_desde debe ser anterior a fecha_hasta'}), 400
        
        # Convertir product_ids sintéticos a (id_variable, id_pais) pairs
        fks_list = []
        for product_id in product_ids:
            id_variable = product_id // 10000
            id_pais = product_id % 10000
            fks_list.append((id_variable, id_pais))
        
        # Construir condiciones WHERE para (id_variable, id_pais) pairs
        fks_conditions = []
        fks_params = []
        for id_var, id_pais in fks_list:
            fks_conditions.append("(m.id_variable = ? AND m.id_pais = ?)")
            fks_params.extend([id_var, id_pais])
        
        # Obtener información de cada cotización/país
        query_products = f"""
            SELECT 
                (m.id_variable * 10000 + m.id_pais) as id,
                v.id_nombre_variable as nombre,
                pg.nombre_pais_grupo as pais,
                m.fuente
            FROM maestro m
            LEFT JOIN variables v ON m.id_variable = v.id_variable
            LEFT JOIN pais_grupo pg ON m.id_pais = pg.id_pais
            WHERE ({' OR '.join(fks_conditions)})
            AND (m.activo = 1 OR CAST(m.activo AS INTEGER) = 1)
        """
        products = execute_query(query_products, tuple(fks_params))
        
        if not products:
            return jsonify({'error': 'No se encontraron cotizaciones activas'}), 404
        
        # Extraer id_pais de los países seleccionados
        id_paises_seleccionados = [id_pais for _, id_pais in fks_list]
        
        # Obtener todos los TC de una vez
        all_tc_monthly = get_all_tc_monthly(id_paises_seleccionados, fecha_desde, fecha_hasta)
        
        # Almacenar datos para las 3 hojas
        all_indices_normalized = {}  # {pais_id: [(fecha, valor), ...]}
        all_indices_original = {}    # {pais_id: [(fecha, valor), ...]}
        all_ipc_data = {}            # {pais_id: [(fecha, valor), ...]}
        all_tc_data = {}             # {pais_id: [(fecha, valor), ...]}
        pais_names = {}               # {pais_id: nombre_pais}
        
        # Procesar cada país
        for product in products:
            product_id = product['id']
            product_name = product['nombre']
            pais = product.get('pais', '')
            id_pais_product = product_id % 10000
            pais_names[id_pais_product] = pais
            
            # Obtener TC mensual del país
            tc_monthly = all_tc_monthly.get(id_pais_product, {})
            if not tc_monthly:
                continue
            
            # Obtener IPC mensual del país
            ipc_monthly = get_ipc_by_country(id_pais_product, fecha_desde, fecha_hasta)
            if not ipc_monthly:
                continue
            
            # Guardar datos originales (IPC y TC)
            all_ipc_data[id_pais_product] = sorted([(fecha, valor) for fecha, valor in ipc_monthly.items()])
            all_tc_data[id_pais_product] = sorted([(fecha, valor) for fecha, valor in tc_monthly.items()])
            
            # Calcular índice original: (IPC / TC)
            indices_orig = []
            for mes_fecha, tc_valor in tc_monthly.items():
                if mes_fecha in ipc_monthly:
                    ipc_valor = ipc_monthly[mes_fecha]
                    if tc_valor > 0:
                        indice_valor = ipc_valor / tc_valor
                        indices_orig.append((mes_fecha, indice_valor))
            
            if len(indices_orig) < 2:
                continue
            
            indices_orig.sort(key=lambda x: x[0])
            all_indices_original[id_pais_product] = indices_orig
            
            # Normalizar a base 100
            first_value = indices_orig[0][1]
            if first_value == 0 or first_value is None:
                continue
            
            factor = 100.0 / first_value
            indices_norm = [(fecha, valor * factor) for fecha, valor in indices_orig]
            all_indices_normalized[id_pais_product] = indices_norm
        
        # Crear Excel
        wb = Workbook()
        
        # Estilos
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_alignment = Alignment(horizontal="center", vertical="center")
        data_alignment = Alignment(horizontal="right", vertical="center")
        
        # Hoja 1: Índice Normalizado
        ws1 = wb.active
        ws1.title = "Índice Normalizado"
        
        ws1['A1'] = 'Fecha'
        ws1['A1'].fill = header_fill
        ws1['A1'].font = header_font
        ws1['A1'].alignment = header_alignment
        
        col = 2
        for pais_id in sorted(all_indices_normalized.keys()):
            cell = ws1.cell(row=1, column=col)
            cell.value = pais_names[pais_id]
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            col += 1
        
        # Obtener todas las fechas únicas
        all_dates = set()
        for indices in all_indices_normalized.values():
            all_dates.update([idx[0] for idx in indices])
        sorted_dates = sorted(all_dates)
        
        # Escribir datos
        row = 2
        for fecha in sorted_dates:
            ws1.cell(row=row, column=1).value = fecha
            col = 2
            for pais_id in sorted(all_indices_normalized.keys()):
                indices = all_indices_normalized[pais_id]
                valor = next((idx[1] for idx in indices if idx[0] == fecha), None)
                if valor is not None:
                    cell = ws1.cell(row=row, column=col)
                    cell.value = valor
                    cell.number_format = '0.00'
                    cell.alignment = data_alignment
                col += 1
            row += 1
        
        # Hoja 2: Índice Original
        ws2 = wb.create_sheet("Índice Original")
        
        ws2['A1'] = 'Fecha'
        ws2['A1'].fill = header_fill
        ws2['A1'].font = header_font
        ws2['A1'].alignment = header_alignment
        
        col = 2
        for pais_id in sorted(all_indices_original.keys()):
            cell = ws2.cell(row=1, column=col)
            cell.value = pais_names[pais_id]
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            col += 1
        
        all_dates_orig = set()
        for indices in all_indices_original.values():
            all_dates_orig.update([idx[0] for idx in indices])
        sorted_dates_orig = sorted(all_dates_orig)
        
        row = 2
        for fecha in sorted_dates_orig:
            ws2.cell(row=row, column=1).value = fecha
            col = 2
            for pais_id in sorted(all_indices_original.keys()):
                indices = all_indices_original[pais_id]
                valor = next((idx[1] for idx in indices if idx[0] == fecha), None)
                if valor is not None:
                    cell = ws2.cell(row=row, column=col)
                    cell.value = valor
                    cell.number_format = '0.00'
                    cell.alignment = data_alignment
                col += 1
            row += 1
        
        # Hoja 3: Datos Originales (IPC y TC)
        ws3 = wb.create_sheet("Datos Originales")
        ws3['A1'] = 'Fecha'
        ws3['A1'].fill = header_fill
        ws3['A1'].font = header_font
        ws3['A1'].alignment = header_alignment
        
        # Encabezados: IPC y TC para cada país
        col = 2
        pais_cols = {}
        for pais_id in sorted(set(list(all_ipc_data.keys()) + list(all_tc_data.keys()))):
            # IPC
            ipc_col = col
            cell = ws3.cell(row=1, column=ipc_col)
            cell.value = f'IPC {pais_names[pais_id]}'
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            col += 1
            
            # TC
            tc_col = col
            cell = ws3.cell(row=1, column=tc_col)
            cell.value = f'TC USD/LC {pais_names[pais_id]}'
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            col += 1
            
            pais_cols[pais_id] = {'ipc': ipc_col, 'tc': tc_col}
        
        # Obtener todas las fechas
        all_dates_raw = set()
        for data in all_ipc_data.values():
            all_dates_raw.update([d[0] for d in data])
        for data in all_tc_data.values():
            all_dates_raw.update([d[0] for d in data])
        sorted_dates_raw = sorted(all_dates_raw)
        
        row = 2
        for fecha in sorted_dates_raw:
            ws3.cell(row=row, column=1).value = fecha
            
            for pais_id in sorted(pais_cols.keys()):
                cols = pais_cols[pais_id]
                
                # IPC
                ipc_valor = next((d[1] for d in all_ipc_data.get(pais_id, []) if d[0] == fecha), None)
                if ipc_valor is not None:
                    cell = ws3.cell(row=row, column=cols['ipc'])
                    cell.value = ipc_valor
                    cell.number_format = '0.00'
                    cell.alignment = data_alignment
                
                # TC
                tc_valor = next((d[1] for d in all_tc_data.get(pais_id, []) if d[0] == fecha), None)
                if tc_valor is not None:
                    cell = ws3.cell(row=row, column=cols['tc'])
                    cell.value = tc_valor
                    cell.number_format = '0.00'
                    cell.alignment = data_alignment
            
            row += 1
        
        # Guardar en memoria
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Generar nombre de archivo
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f'inflacion_dolares_{timestamp}.xlsx'
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
    
    except ValueError as e:
        return jsonify({'error': f'Error en formato de fecha: {str(e)}'}), 400
    except Exception as e:
        print(f"[ERROR] export_inflacion_dolares_to_excel: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': f'Error al generar Excel: {str(e)}'}), 500
