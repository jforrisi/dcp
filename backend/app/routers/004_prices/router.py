"""API routes for price data."""
import importlib
from datetime import date, datetime
from typing import List, Optional
from io import BytesIO
from flask import Blueprint, request, jsonify, abort, send_file
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from ...database import execute_query, execute_query_single

# Import from numbered module using importlib
_dcp_module = importlib.import_module('app.routers.001_dcp.router')
get_macro_series = _dcp_module.get_macro_series
convert_to_monthly = _dcp_module.convert_to_monthly
get_product_currency = _dcp_module.get_product_currency
TC_USD_ID = _dcp_module.TC_USD_ID
TC_EUR_ID = _dcp_module.TC_EUR_ID
IPC_ID = _dcp_module.IPC_ID

bp = Blueprint('prices', __name__)


@bp.route('/products', methods=['GET'])
def get_products():
    """
    Get all active products/services para Precios Corrientes.
    Usa los mismos filtros que DCP:
    - id_familia = 2 (todas las subfamilias, sin restricciones)
    - O id_familia = 3 con id_subfamilia IN (5, 4, 3, 2) 
      Y id_pais = 858 (Uruguay) 
      Y id_variable != 9 (excluir IPC)
    """
    # Usar LEFT JOINs para obtener información de nuevas tablas si FKs existen
    # NOTA: Ya no hay columna 'id', 'nombre', 'tipo', 'categoria', 'moneda', 'nominal_real'
    # Generamos un id sintético para compatibilidad: id_variable * 10000 + id_pais
    query = """
        SELECT 
            (m.id_variable * 10000 + m.id_pais) as id,
            v.id_nombre_variable as nombre,
            m.fuente,
            m.periodicidad,
            m.activo,
            v.moneda,
            v.nominal_o_real as nominal_real,
            pg.nombre_pais_grupo as pais,
            sf.nombre_sub_familia as sub_familia,
            f.nombre_familia as familia,
            m.id_variable,
            m.id_pais
        FROM maestro m
        LEFT JOIN variables v ON m.id_variable = v.id_variable
        LEFT JOIN pais_grupo pg ON m.id_pais = pg.id_pais
        LEFT JOIN sub_familia sf ON v.id_sub_familia = sf.id_sub_familia
        LEFT JOIN familia f ON sf.id_familia = f.id_familia
        WHERE m.activo = 1
        AND (
            sf.id_familia = 2
            OR (sf.id_familia = 3 
                AND sf.id_sub_familia IN (5, 4, 3, 2)
                AND m.id_pais = 858
                AND m.id_variable != 9)
        )
        ORDER BY v.id_nombre_variable
    """
    results = execute_query(query)
    return jsonify(results)


@bp.route('/products/<int:product_id>/prices', methods=['GET'])
def get_product_prices(product_id: int):
    """Get prices for a single product within a date range."""
    # Convertir id sintético a id_variable e id_pais
    # id sintético = id_variable * 10000 + id_pais
    id_variable = product_id // 10000
    id_pais = product_id % 10000
    
    # Verificar que existe en maestro
    query_check = """
        SELECT id_variable, id_pais FROM maestro WHERE id_variable = ? AND id_pais = ?
    """
    check_result = execute_query_single(query_check, (id_variable, id_pais))
    
    if not check_result:
        return jsonify([])
    
    fecha_desde = request.args.get('fecha_desde', type=str)
    fecha_hasta = request.args.get('fecha_hasta', type=str)
    
    # Convert string dates to date objects if provided
    if fecha_desde:
        fecha_desde = date.fromisoformat(fecha_desde)
    if fecha_hasta:
        fecha_hasta = date.fromisoformat(fecha_hasta)
    
    if fecha_desde and fecha_hasta:
        query = """
            SELECT fecha, valor
            FROM maestro_precios
            WHERE id_variable = ? AND id_pais = ? AND fecha BETWEEN ? AND ?
            ORDER BY fecha ASC
        """
        params = (id_variable, id_pais, fecha_desde, fecha_hasta)
    elif fecha_desde:
        query = """
            SELECT fecha, valor
            FROM maestro_precios
            WHERE id_variable = ? AND id_pais = ? AND fecha >= ?
            ORDER BY fecha ASC
        """
        params = (id_variable, id_pais, fecha_desde)
    elif fecha_hasta:
        query = """
            SELECT fecha, valor
            FROM maestro_precios
            WHERE id_variable = ? AND id_pais = ? AND fecha <= ?
            ORDER BY fecha ASC
        """
        params = (id_variable, id_pais, fecha_hasta)
    else:
        query = """
            SELECT fecha, valor
            FROM maestro_precios
            WHERE id_variable = ? AND id_pais = ?
            ORDER BY fecha ASC
        """
        params = (id_variable, id_pais)

    results = execute_query(query, params)
    return jsonify(results)


@bp.route('/products/prices', methods=['GET'])
def get_multiple_products_prices():
    """Get prices for multiple products within a date range."""
    # Get product_ids from query params (can be multiple with same key)
    product_ids = request.args.getlist('product_ids[]', type=int)
    
    if not product_ids:
        return jsonify([])
    
    fecha_desde = request.args.get('fecha_desde', type=str)
    fecha_hasta = request.args.get('fecha_hasta', type=str)
    
    # Convert string dates to date objects if provided
    if fecha_desde:
        fecha_desde = date.fromisoformat(fecha_desde)
    if fecha_hasta:
        fecha_hasta = date.fromisoformat(fecha_hasta)

    # Convertir product_ids sintéticos a (id_variable, id_pais)
    # id sintético = id_variable * 10000 + id_pais
    fks_map = {}
    for product_id in product_ids:
        id_variable = product_id // 10000
        id_pais = product_id % 10000
        # Verificar que existe en maestro
        query_check = """
            SELECT id_variable, id_pais FROM maestro WHERE id_variable = ? AND id_pais = ?
        """
        check_result = execute_query_single(query_check, (id_variable, id_pais))
        if check_result:
            fks_map[product_id] = (id_variable, id_pais)
    
    if not fks_map:
        # Si ningún producto tiene FKs, retornar vacío
        return jsonify([])
    
    # Construir condiciones WHERE para (id_variable, id_pais) pairs
    fks_conditions = []
    fks_params = []
    for id_var, id_pais in fks_map.values():
        fks_conditions.append("(mp.id_variable = ? AND mp.id_pais = ?)")
        fks_params.extend([id_var, id_pais])
    
    fks_where = " OR ".join(fks_conditions)
    
    # Build query with fecha conditions
    if fecha_desde and fecha_hasta:
        query = f"""
            SELECT mp.id_variable, mp.id_pais, mp.fecha, mp.valor,
                   (m.id_variable * 10000 + m.id_pais) as id,
                   v.id_nombre_variable as nombre,
                   m.fuente, m.periodicidad
            FROM maestro_precios mp
            JOIN maestro m ON mp.id_variable = m.id_variable AND mp.id_pais = m.id_pais
            LEFT JOIN variables v ON m.id_variable = v.id_variable
            WHERE ({fks_where}) 
              AND mp.fecha BETWEEN ? AND ?
            ORDER BY (m.id_variable * 10000 + m.id_pais), mp.fecha ASC
        """
        params = tuple(fks_params) + (fecha_desde, fecha_hasta)
    elif fecha_desde:
        query = f"""
            SELECT mp.id_variable, mp.id_pais, mp.fecha, mp.valor,
                   (m.id_variable * 10000 + m.id_pais) as id,
                   v.id_nombre_variable as nombre,
                   m.fuente, m.periodicidad
            FROM maestro_precios mp
            JOIN maestro m ON mp.id_variable = m.id_variable AND mp.id_pais = m.id_pais
            LEFT JOIN variables v ON m.id_variable = v.id_variable
            WHERE ({fks_where}) 
              AND mp.fecha >= ?
            ORDER BY (m.id_variable * 10000 + m.id_pais), mp.fecha ASC
        """
        params = tuple(fks_params) + (fecha_desde,)
    elif fecha_hasta:
        query = f"""
            SELECT mp.id_variable, mp.id_pais, mp.fecha, mp.valor,
                   (m.id_variable * 10000 + m.id_pais) as id,
                   v.id_nombre_variable as nombre,
                   m.fuente, m.periodicidad
            FROM maestro_precios mp
            JOIN maestro m ON mp.id_variable = m.id_variable AND mp.id_pais = m.id_pais
            LEFT JOIN variables v ON m.id_variable = v.id_variable
            WHERE ({fks_where}) 
              AND mp.fecha <= ?
            ORDER BY (m.id_variable * 10000 + m.id_pais), mp.fecha ASC
        """
        params = tuple(fks_params) + (fecha_hasta,)
    else:
        query = f"""
            SELECT mp.id_variable, mp.id_pais, mp.fecha, mp.valor,
                   (m.id_variable * 10000 + m.id_pais) as id,
                   v.id_nombre_variable as nombre,
                   m.fuente, m.periodicidad
            FROM maestro_precios mp
            JOIN maestro m ON mp.id_variable = m.id_variable AND mp.id_pais = m.id_pais
            LEFT JOIN variables v ON m.id_variable = v.id_variable
            WHERE ({fks_where})
            ORDER BY (m.id_variable * 10000 + m.id_pais), mp.fecha ASC
        """
        params = tuple(fks_params)

    results = execute_query(query, params)
    
    # Group by product and format response
    products_dict = {}
    for row in results:
        product_id = row['id']  # Synthetic ID: id_variable * 10000 + id_pais
        if product_id not in products_dict:
            products_dict[product_id] = {
                'product_id': row['id'],
                'product_name': row['nombre'],
                'unit': None,  # unidad no existe en nuevo schema
                'data': []
            }
        products_dict[product_id]['data'].append({
            'fecha': row['fecha'],
            'valor': row['valor']
        })
    
    # Calcular summary para cada producto
    result_list = []
    for product_id, product_data in products_dict.items():
        data = product_data['data']
        if data:
            # Ordenar por fecha
            data.sort(key=lambda x: x['fecha'])
            precio_inicial = data[0]['valor']
            precio_final = data[-1]['valor']
            variacion = 0.0
            if precio_inicial > 0:
                variacion = ((precio_final - precio_inicial) / precio_inicial) * 100
            
            fecha_inicial = data[0]['fecha']
            fecha_final = data[-1]['fecha']
            
            product_data['summary'] = {
                'precio_inicial': precio_inicial,
                'precio_final': precio_final,
                'variacion_nominal': variacion,
                'fecha_inicial': fecha_inicial,
                'fecha_final': fecha_final
            }
        else:
            product_data['summary'] = None
        
        result_list.append(product_data)
    
    return jsonify(result_list)


@bp.route('/variations', methods=['GET'])
def get_price_variations():
    """Calculate DCP index variations for all products in a date range."""
    fecha_desde = request.args.get('fecha_desde', type=str)
    fecha_hasta = request.args.get('fecha_hasta', type=str)
    order_by = request.args.get('order_by', default='desc', type=str)
    
    if not fecha_desde or not fecha_hasta:
        abort(400, description="fecha_desde and fecha_hasta are required")
    
    if order_by not in ['asc', 'desc']:
        order_by = 'desc'
    
    # Convert string dates to date objects
    try:
        fecha_desde = date.fromisoformat(fecha_desde)
        fecha_hasta = date.fromisoformat(fecha_hasta)
    except ValueError as e:
        abort(400, description=f"Invalid date format: {str(e)}")
    
    # Obtener todos los productos y servicios activos (incluye nominal_real/moneda)
    # Usar LEFT JOINs para obtener moneda y nominal_real desde variables si FKs existen
    query_products = """
        SELECT 
            (m.id_variable * 10000 + m.id_pais) as id,
            v.id_nombre_variable as nombre,
            m.periodicidad,
            v.moneda as moneda,
            v.nominal_o_real as nominal_real
        FROM maestro m
        LEFT JOIN variables v ON m.id_variable = v.id_variable
        WHERE m.activo = 1
        ORDER BY v.id_nombre_variable
    """
    products = execute_query(query_products)
    
    if not products:
        return jsonify({'variations': [], 'omitted_products': []})
    
    # Obtener TC e IPC mensuales
    tc_usd_monthly = get_macro_series(TC_USD_ID, fecha_desde, fecha_hasta)
    tc_eur_monthly = get_macro_series(TC_EUR_ID, fecha_desde, fecha_hasta)
    ipc_monthly = get_macro_series(IPC_ID, fecha_desde, fecha_hasta)
    
    if not ipc_monthly:
        abort(400, description="IPC data not available for the selected date range")
    
    result = []
    omitted_products = []
    
    # Procesar cada producto
    for product in products:
        product_id = product['id']
        product_name = product['nombre']
        periodicidad = product['periodicidad']
        unidad = None  # unidad no existe en nuevo schema
        
        # Convertir id sintético a id_variable e id_pais
        id_variable = product_id // 10000
        id_pais = product_id % 10000
        
        # Verificar que existe en maestro
        query_check = "SELECT id_variable, id_pais FROM maestro WHERE id_variable = ? AND id_pais = ?"
        fks_result = execute_query_single(query_check, (id_variable, id_pais))
        
        if not fks_result:
            omitted_products.append({
                'id': product_id,
                'nombre': product_name,
                'razon': "No tiene id_variable e id_pais",
                'fecha_ultimo_dato': None
            })
            continue
        
        # id_variable e id_pais ya están calculados arriba
        
        # Obtener precios del producto (ampliar rango para encontrar última fecha disponible)
        query_prices = """
            SELECT fecha, valor
            FROM maestro_precios
            WHERE id_variable = ? AND id_pais = ? AND fecha <= ?
            ORDER BY fecha ASC
        """
        try:
            raw_prices = execute_query(query_prices, (id_variable, id_pais, fecha_hasta))
        except Exception as e:
            print(f"[VARIATIONS] Error obteniendo precios para producto {product_id} ({product_name}): {e}")
            omitted_products.append({
                'id': product_id,
                'nombre': product_name,
                'razon': f"Error al obtener precios: {str(e)}",
                'fecha_ultimo_dato': None
            })
            continue
        
        if not raw_prices:
            # Buscar última fecha disponible fuera del rango
            query_last_date = """
                SELECT MAX(fecha) as ultima_fecha
                FROM maestro_precios
                WHERE id_variable = ? AND id_pais = ?
            """
            try:
                last_date_result = execute_query_single(query_last_date, (id_variable, id_pais))
                last_date = last_date_result['ultima_fecha'] if last_date_result and last_date_result['ultima_fecha'] else None
            except:
                last_date = None
            
            omitted_products.append({
                'id': product_id,
                'nombre': product_name,
                'razon': "No hay datos de precios en el rango seleccionado",
                'fecha_ultimo_dato': str(last_date) if last_date else None
            })
            continue
        
        # Convertir precios a mensual
        prices_monthly = convert_to_monthly(raw_prices, periodicidad)
        
        # Determinar qué tipo de cambio usar basándose en la moneda del producto
        moneda = (product.get('moneda') or get_product_currency(product_id))
        nominal_real = (product.get('nominal_real') or 'n').lower()
        if moneda == 'eur':
            tc_monthly = tc_eur_monthly
        elif moneda == 'usd':
            tc_monthly = tc_usd_monthly
        elif moneda == 'uyu' or moneda is None:
            # Para productos en UYU o sin moneda definida, usar TC = 1.0
            tc_monthly = {fecha: 1.0 for fecha in ipc_monthly.keys()}
        else:
            # Moneda desconocida, usar USD por defecto
            tc_monthly = tc_usd_monthly
        
        if not tc_monthly:
            if moneda == 'eur':
                tc_type = "EUR/UYU"
            elif moneda == 'usd':
                tc_type = "USD/UYU"
            else:
                tc_type = "USD/UYU"
            omitted_products.append({
                'id': product_id,
                'nombre': product_name,
                'razon': f"No hay datos de tipo de cambio {tc_type} disponibles en el rango seleccionado",
                'fecha_ultimo_dato': None
            })
            continue
        
        # Calcular índices con nominal/real:
        # base = precio × TC; si nominal_real == 'n' -> dividir por IPC; si 'r' -> no dividir
        indices = []
        for price_item in prices_monthly:
            mes_fecha = price_item['fecha']
            precio = float(price_item['valor'])
            
            # Asegurar que mes_fecha sea un objeto date
            if not isinstance(mes_fecha, date):
                if isinstance(mes_fecha, str):
                    mes_fecha = date.fromisoformat(mes_fecha)
                else:
                    continue
            
            # Verificar que exista TC (y si es nominal, IPC) para este mes
            if mes_fecha in tc_monthly:
                tc_valor = float(tc_monthly[mes_fecha])
                base_valor = precio * tc_valor
                
                if nominal_real == 'r':
                    indices.append({'fecha': mes_fecha, 'valor': base_valor})
                else:
                    if mes_fecha in ipc_monthly:
                        ipc_valor = float(ipc_monthly[mes_fecha])
                        if ipc_valor > 0:
                            indices.append({'fecha': mes_fecha, 'valor': base_valor / ipc_valor})
        
        if len(indices) < 2:
            # Buscar última fecha con índice calculado
            last_index_date = max([idx['fecha'] for idx in indices]) if indices else None
            omitted_products.append({
                'id': product_id,
                'nombre': product_name,
                'razon': "Menos de 2 índices calculados (faltan datos de IPC o TC para algunos meses)",
                'fecha_ultimo_dato': str(last_index_date) if last_index_date else None
            })
            continue
        
        # Filtrar índices dentro del rango de fechas (igual que en DCP)
        indices_filtered = [idx for idx in indices if idx['fecha'] >= fecha_desde and idx['fecha'] <= fecha_hasta]
        
        if len(indices_filtered) < 2:
            # Buscar última fecha con índice en el rango
            last_index_date = max([idx['fecha'] for idx in indices_filtered]) if indices_filtered else None
            if not last_index_date:
                # Usar última fecha de todos los índices calculados
                last_index_date = max([idx['fecha'] for idx in indices]) if indices else None
            
            omitted_products.append({
                'id': product_id,
                'nombre': product_name,
                'razon': f"Menos de 2 índices dentro del rango seleccionado",
                'fecha_ultimo_dato': str(last_index_date) if last_index_date else None
            })
            continue
        
        # Ordenar por fecha
        indices_filtered.sort(key=lambda x: x['fecha'])
        
        # Obtener primer y último índice del rango filtrado
        indice_inicial = indices_filtered[0]['valor']
        indice_final = indices_filtered[-1]['valor']
        fecha_inicial = indices_filtered[0]['fecha']
        fecha_final = indices_filtered[-1]['fecha']
        
        if indice_inicial == 0 or indice_inicial is None:
            omitted_products.append({
                'id': product_id,
                'nombre': product_name,
                'razon': "Índice inicial es cero o nulo, no se puede calcular variación",
                'fecha_ultimo_dato': str(fecha_inicial) if fecha_inicial else None
            })
            continue
        
        # Calcular variación porcentual
        variacion_percent = ((indice_final - indice_inicial) / indice_inicial) * 100.0
        
        # Log para debugging
        print(f"[VARIATIONS] Producto {product_id} ({product_name}): "
              f"Rango {fecha_inicial} a {fecha_final}, "
              f"Índice inicial: {indice_inicial:.2f}, final: {indice_final:.2f}, "
              f"Variación: {variacion_percent:.2f}%")
        
        result.append({
            'id': product_id,
            'nombre': product_name,
            'unidad': unidad,
            'precio_inicial': indice_inicial,  # Mantener nombre para compatibilidad con frontend
            'precio_final': indice_final,  # Mantener nombre para compatibilidad con frontend
            'variacion_percent': variacion_percent
        })
    
    # Ordenar resultados
    result.sort(key=lambda x: x['variacion_percent'], reverse=(order_by == 'desc'))
    
    # Log para debugging
    print(f"[VARIATIONS] Total productos procesados: {len(products)}")
    print(f"[VARIATIONS] Productos con variación calculada: {len(result)}")
    print(f"[VARIATIONS] Productos omitidos: {len(omitted_products)}")
    for omitted in omitted_products:
        print(f"[VARIATIONS] Omitido - {omitted['nombre']}: {omitted['razon']}")
    
    return jsonify({
        'variations': result,
        'omitted_products': omitted_products
    })


@bp.route('/variations/export', methods=['GET'])
def export_variations_to_excel():
    """Export variations data to Excel with complete calculation details."""
    fecha_desde_str = request.args.get('fecha_desde', type=str)
    fecha_hasta_str = request.args.get('fecha_hasta', type=str)
    order_by = request.args.get('order_by', default='desc', type=str)
    
    if not fecha_desde_str or not fecha_hasta_str:
        abort(400, description="fecha_desde and fecha_hasta are required")
    
    try:
        fecha_desde = date.fromisoformat(fecha_desde_str)
        fecha_hasta = date.fromisoformat(fecha_hasta_str)
    except ValueError as e:
        abort(400, description=f"Invalid date format: {str(e)}")
    
    # Obtener todos los productos y servicios activos (incluye nominal_real/moneda)
    # Usar LEFT JOINs para obtener moneda y nominal_real desde variables si FKs existen
    query_products = """
        SELECT 
            (m.id_variable * 10000 + m.id_pais) as id,
            v.id_nombre_variable as nombre,
            m.periodicidad,
            v.moneda as moneda,
            v.nominal_o_real as nominal_real
        FROM maestro m
        LEFT JOIN variables v ON m.id_variable = v.id_variable
        WHERE m.activo = 1
        ORDER BY v.id_nombre_variable
    """
    products = execute_query(query_products)
    
    if not products:
        abort(400, description="No products found")
    
    # Obtener TC e IPC mensuales
    tc_usd_monthly = get_macro_series(TC_USD_ID, fecha_desde, fecha_hasta)
    tc_eur_monthly = get_macro_series(TC_EUR_ID, fecha_desde, fecha_hasta)
    ipc_monthly = get_macro_series(IPC_ID, fecha_desde, fecha_hasta)
    
    if not ipc_monthly:
        abort(400, description="IPC data not available for the selected date range")
    
    # Estilos
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_alignment = Alignment(horizontal="center", vertical="center")
    data_alignment = Alignment(horizontal="right", vertical="center")
    
    wb = Workbook()
    
    # Datos para las hojas
    summary_data = []  # Resumen variaciones
    all_indices_calculated = {}  # {product_id: [(fecha, indice), ...]}
    all_indices_original = {}  # {product_id: [(fecha, indice), ...]}
    all_prices_original = {}  # {product_id: [(fecha, precio), ...]}
    product_names = {}  # {product_id: nombre}
    omitted_products = []
    
    # Procesar cada producto (similar a /variations pero guardando más datos)
    for product in products:
        product_id = product['id']
        product_name = product['nombre']
        periodicidad = product['periodicidad']
        unidad = None  # unidad no existe en nuevo schema
        product_names[product_id] = product_name
        
        # Convertir id sintético a id_variable e id_pais
        id_variable = product_id // 10000
        id_pais = product_id % 10000
        
        # Verificar que existe en maestro
        query_check = "SELECT id_variable, id_pais FROM maestro WHERE id_variable = ? AND id_pais = ?"
        fks_result = execute_query_single(query_check, (id_variable, id_pais))
        
        if not fks_result:
            omitted_products.append({
                'id': product_id,
                'nombre': product_name,
                'razon': "No tiene id_variable e id_pais"
            })
            continue
        
        # id_variable e id_pais ya están calculados arriba
        
        # Obtener precios hasta fecha_hasta inclusive (para mostrar datos hasta donde hay disponibles)
        query_prices = """
            SELECT fecha, valor
            FROM maestro_precios
            WHERE id_variable = ? AND id_pais = ? AND fecha <= ?
            ORDER BY fecha ASC
        """
        try:
            raw_prices = execute_query(query_prices, (id_variable, id_pais, fecha_hasta))
        except Exception as e:
            omitted_products.append({
                'id': product_id,
                'nombre': product_name,
                'razon': f"Error al obtener precios: {str(e)}"
            })
            continue
        
        if not raw_prices:
            omitted_products.append({
                'id': product_id,
                'nombre': product_name,
                'razon': "No hay datos de precios en el rango seleccionado"
            })
            continue
        
        # Convertir precios a mensual
        prices_monthly = convert_to_monthly(raw_prices, periodicidad)
        
        # Guardar precios mensuales (estos son los que se usan para calcular índices)
        for price_item in prices_monthly:
            mes_fecha = price_item['fecha']
            precio = float(price_item['valor'])
            
            if not isinstance(mes_fecha, date):
                if isinstance(mes_fecha, str):
                    mes_fecha = date.fromisoformat(mes_fecha)
                else:
                    continue
            
            if product_id not in all_prices_original:
                all_prices_original[product_id] = []
            all_prices_original[product_id].append((mes_fecha, precio))
        
        # Determinar qué tipo de cambio usar basándose en la moneda del producto
        moneda = (product.get('moneda') or get_product_currency(product_id))
        nominal_real = (product.get('nominal_real') or 'n').lower()
        if moneda == 'eur':
            tc_monthly = tc_eur_monthly
        elif moneda == 'usd':
            tc_monthly = tc_usd_monthly
        elif moneda == 'uyu' or moneda is None:
            # Para productos en UYU o sin moneda definida, usar TC = 1.0
            tc_monthly = {fecha: 1.0 for fecha in ipc_monthly.keys()}
        else:
            # Moneda desconocida, usar USD por defecto
            tc_monthly = tc_usd_monthly
        
        if not tc_monthly:
            if moneda == 'eur':
                tc_type = "EUR/UYU"
            elif moneda == 'usd':
                tc_type = "USD/UYU"
            else:
                tc_type = "USD/UYU"
            omitted_products.append({
                'id': product_id,
                'nombre': product_name,
                'razon': f"No hay datos de tipo de cambio {tc_type} disponibles"
            })
            continue
        
        # Calcular índices con nominal/real:
        # base = precio × TC; si nominal_real == 'n' -> dividir por IPC; si 'r' -> no dividir
        indices = []
        indices_original = []
        for price_item in prices_monthly:
            mes_fecha = price_item['fecha']
            precio = float(price_item['valor'])
            
            if not isinstance(mes_fecha, date):
                if isinstance(mes_fecha, str):
                    mes_fecha = date.fromisoformat(mes_fecha)
                else:
                    continue
            
            if mes_fecha in tc_monthly:
                tc_valor = float(tc_monthly[mes_fecha])
                base_valor = precio * tc_valor
                if nominal_real == 'r':
                    indices.append({'fecha': mes_fecha, 'valor': base_valor})
                    indices_original.append((mes_fecha, base_valor))
                else:
                    if mes_fecha in ipc_monthly:
                        ipc_valor = float(ipc_monthly[mes_fecha])
                        if ipc_valor > 0:
                            indice = base_valor / ipc_valor
                            indices.append({'fecha': mes_fecha, 'valor': indice})
                            indices_original.append((mes_fecha, indice))
        
        if len(indices) < 2:
            omitted_products.append({
                'id': product_id,
                'nombre': product_name,
                'razon': "Menos de 2 índices calculados"
            })
            continue
        
        # Filtrar índices dentro del rango
        indices_filtered = [idx for idx in indices if idx['fecha'] >= fecha_desde and idx['fecha'] <= fecha_hasta]
        
        if len(indices_filtered) < 2:
            omitted_products.append({
                'id': product_id,
                'nombre': product_name,
                'razon': "Menos de 2 índices dentro del rango seleccionado"
            })
            continue
        
        # Ordenar por fecha
        indices_filtered.sort(key=lambda x: x['fecha'])
        
        # Obtener primer y último índice
        indice_inicial = indices_filtered[0]['valor']
        indice_final = indices_filtered[-1]['valor']
        fecha_inicial = indices_filtered[0]['fecha']
        fecha_final = indices_filtered[-1]['fecha']
        
        if indice_inicial == 0 or indice_inicial is None:
            omitted_products.append({
                'id': product_id,
                'nombre': product_name,
                'razon': "Índice inicial es cero o nulo"
            })
            continue
        
        # Calcular variación
        variacion_percent = ((indice_final - indice_inicial) / indice_inicial) * 100.0
        
        # Guardar datos para Excel
        summary_data.append({
            'product_id': product_id,
            'nombre': product_name,
            'unidad': unidad,
            'variacion_percent': variacion_percent,
            'indice_inicial': indice_inicial,
            'indice_final': indice_final,
            'fecha_inicial': fecha_inicial,
            'fecha_final': fecha_final
        })
        
        # Guardar índices calculados (filtrados)
        all_indices_calculated[product_id] = [(idx['fecha'], idx['valor']) for idx in indices_filtered]
        
        # Guardar índices originales (todos los calculados)
        all_indices_original[product_id] = indices_original
    
    # Ordenar resumen
    summary_data.sort(key=lambda x: x['variacion_percent'], reverse=(order_by == 'desc'))
    
    # Hoja 1: Resumen Variaciones
    ws1 = wb.active
    ws1.title = "Resumen Variaciones"
    
    headers1 = ['Producto', 'Unidad', 'Variación %', 'Índice Inicial', 'Índice Final', 'Fecha Inicial', 'Fecha Final']
    for col, header in enumerate(headers1, 1):
        cell = ws1.cell(row=1, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    for row_idx, item in enumerate(summary_data, 2):
        ws1.cell(row=row_idx, column=1).value = item['nombre']
        ws1.cell(row=row_idx, column=2).value = item['unidad'] or ''
        ws1.cell(row=row_idx, column=3).value = item['variacion_percent']
        ws1.cell(row=row_idx, column=3).number_format = '0.00'
        ws1.cell(row=row_idx, column=4).value = item['indice_inicial']
        ws1.cell(row=row_idx, column=4).number_format = '0.00'
        ws1.cell(row=row_idx, column=5).value = item['indice_final']
        ws1.cell(row=row_idx, column=5).number_format = '0.00'
        ws1.cell(row=row_idx, column=6).value = item['fecha_inicial']
        ws1.cell(row=row_idx, column=7).value = item['fecha_final']
        ws1.cell(row=row_idx, column=4).alignment = data_alignment
        ws1.cell(row=row_idx, column=5).alignment = data_alignment
        ws1.cell(row=row_idx, column=3).alignment = data_alignment
    
    # Hoja 2: Índices Calculados (filtrados)
    ws2 = wb.create_sheet("Índices Calculados")
    ws2['A1'] = 'Fecha'
    ws2['A1'].fill = header_fill
    ws2['A1'].font = header_font
    ws2['A1'].alignment = header_alignment
    
    col = 2
    for product_id in sorted(all_indices_calculated.keys()):
        cell = ws2.cell(row=1, column=col)
        cell.value = product_names[product_id]
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        col += 1
    
    all_dates_calc = set()
    for indices in all_indices_calculated.values():
        all_dates_calc.update([idx[0] for idx in indices])
    sorted_dates_calc = sorted(all_dates_calc)
    
    row = 2
    for fecha in sorted_dates_calc:
        ws2.cell(row=row, column=1).value = fecha
        col = 2
        for product_id in sorted(all_indices_calculated.keys()):
            indices = all_indices_calculated[product_id]
            valor = next((idx[1] for idx in indices if idx[0] == fecha), None)
            if valor is not None:
                cell = ws2.cell(row=row, column=col)
                cell.value = valor
                cell.number_format = '0.00'
                cell.alignment = data_alignment
            col += 1
        row += 1
    
    # Hoja 3: Índices Originales (todos los calculados)
    ws3 = wb.create_sheet("Índices Originales")
    ws3['A1'] = 'Fecha'
    ws3['A1'].fill = header_fill
    ws3['A1'].font = header_font
    ws3['A1'].alignment = header_alignment
    
    col = 2
    for product_id in sorted(all_indices_original.keys()):
        cell = ws3.cell(row=1, column=col)
        cell.value = product_names[product_id]
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        col += 1
    
    all_dates_orig = set()
    for indices in all_indices_original.values():
        all_dates_orig.update([idx[0] for idx in indices])
    sorted_dates_orig = sorted(all_dates_orig)
    
    row = 2
    for fecha in sorted_dates_orig:
        ws3.cell(row=row, column=1).value = fecha
        col = 2
        for product_id in sorted(all_indices_original.keys()):
            indices = all_indices_original[product_id]
            valor = next((idx[1] for idx in indices if idx[0] == fecha), None)
            if valor is not None:
                cell = ws3.cell(row=row, column=col)
                cell.value = valor
                cell.number_format = '0.00'
                cell.alignment = data_alignment
            col += 1
        row += 1
    
    # Hoja 4: Precios Originales (con componentes: productos, IPC, TC USD/UYU, TC EUR/UYU)
    ws4 = wb.create_sheet("Precios Originales")
    ws4['A1'] = 'Fecha'
    ws4['A1'].fill = header_fill
    ws4['A1'].font = header_font
    ws4['A1'].alignment = header_alignment
    
    # Encabezados: primero productos, luego IPC, TC USD/UYU, TC EUR/UYU
    col = 2
    for product_id in sorted(all_prices_original.keys()):
        cell = ws4.cell(row=1, column=col)
        cell.value = product_names[product_id]
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        col += 1
    
    # Agregar IPC
    ws4.cell(row=1, column=col).value = 'IPC'
    ws4.cell(row=1, column=col).fill = header_fill
    ws4.cell(row=1, column=col).font = header_font
    ws4.cell(row=1, column=col).alignment = header_alignment
    ipc_col = col
    col += 1
    
    # Agregar TC USD/UYU
    ws4.cell(row=1, column=col).value = 'TC USD/UYU'
    ws4.cell(row=1, column=col).fill = header_fill
    ws4.cell(row=1, column=col).font = header_font
    ws4.cell(row=1, column=col).alignment = header_alignment
    tc_usd_col = col
    col += 1
    
    # Agregar TC EUR/UYU
    ws4.cell(row=1, column=col).value = 'TC EUR/UYU'
    ws4.cell(row=1, column=col).fill = header_fill
    ws4.cell(row=1, column=col).font = header_font
    ws4.cell(row=1, column=col).alignment = header_alignment
    tc_eur_col = col
    
    # Obtener todas las fechas (de precios, IPC y TC)
    all_dates_prices = set()
    for prices in all_prices_original.values():
        all_dates_prices.update([p[0] for p in prices])
    all_dates_prices.update(ipc_monthly.keys())
    all_dates_prices.update(tc_usd_monthly.keys())
    all_dates_prices.update(tc_eur_monthly.keys())
    sorted_dates_prices = sorted(all_dates_prices)
    
    row = 2
    for fecha in sorted_dates_prices:
        ws4.cell(row=row, column=1).value = fecha
        
        # Escribir precios de productos
        col = 2
        for product_id in sorted(all_prices_original.keys()):
            prices = all_prices_original[product_id]
            valor = next((p[1] for p in prices if p[0] == fecha), None)
            if valor is not None:
                cell = ws4.cell(row=row, column=col)
                cell.value = valor
                cell.number_format = '0.00'
                cell.alignment = data_alignment
            col += 1
        
        # Escribir IPC
        ipc_valor = ipc_monthly.get(fecha)
        if ipc_valor is not None:
            cell = ws4.cell(row=row, column=ipc_col)
            cell.value = ipc_valor
            cell.number_format = '0.00'
            cell.alignment = data_alignment
        
        # Escribir TC USD/UYU
        tc_usd_valor = tc_usd_monthly.get(fecha)
        if tc_usd_valor is not None:
            cell = ws4.cell(row=row, column=tc_usd_col)
            cell.value = tc_usd_valor
            cell.number_format = '0.00'
            cell.alignment = data_alignment
        
        # Escribir TC EUR/UYU
        tc_eur_valor = tc_eur_monthly.get(fecha)
        if tc_eur_valor is not None:
            cell = ws4.cell(row=row, column=tc_eur_col)
            cell.value = tc_eur_valor
            cell.number_format = '0.00'
            cell.alignment = data_alignment
        
        row += 1
    
    # Hoja 5: Metadatos
    ws5 = wb.create_sheet("Metadatos")
    ws5['A1'] = 'Campo'
    ws5['A1'].fill = header_fill
    ws5['A1'].font = header_font
    ws5['B1'] = 'Valor'
    ws5['B1'].fill = header_fill
    ws5['B1'].font = header_font
    
    metadata = [
        ('Fecha de exportación', datetime.now().strftime('%Y-%m-%d %H:%M:%S')),
        ('Rango de fechas', f'{fecha_desde} a {fecha_hasta}'),
        ('Productos incluidos', ', '.join([product_names[pid] for pid in sorted(all_indices_calculated.keys())])),
        ('Productos omitidos', ', '.join([p['nombre'] for p in omitted_products]) if omitted_products else 'Ninguno'),
        ('Fórmula aplicada', 'Precio internacional × TC / IPC'),
    ]
    
    row = 2
    for campo, valor in metadata:
        ws5.cell(row=row, column=1).value = campo
        ws5.cell(row=row, column=2).value = valor
        row += 1
    
    # Agregar razones de omisión
    if omitted_products:
        ws5.cell(row=row, column=1).value = 'Razones de omisión'
        ws5.cell(row=row, column=1).font = Font(bold=True)
        row += 1
        for omitted in omitted_products:
            ws5.cell(row=row, column=1).value = omitted['nombre']
            ws5.cell(row=row, column=2).value = omitted['razon']
            row += 1
    
    # Ajustar anchos de columna
    for ws in [ws1, ws2, ws3]:
        ws.column_dimensions['A'].width = 15
        for col in range(2, len(all_indices_calculated) + 2):
            ws.column_dimensions[chr(64 + col)].width = 25
    
    # Para ws4 (Precios Originales), ajustar considerando productos + IPC + TC USD + TC EUR
    ws4.column_dimensions['A'].width = 15
    num_product_cols = len(all_prices_original)
    for col in range(2, num_product_cols + 2):
        ws4.column_dimensions[chr(64 + col)].width = 25
    # IPC, TC USD/UYU, TC EUR/UYU
    for col in range(num_product_cols + 2, num_product_cols + 5):
        ws4.column_dimensions[chr(64 + col)].width = 15
    
    ws5.column_dimensions['A'].width = 30
    ws5.column_dimensions['B'].width = 50
    
    # Guardar a BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Nombre de archivo
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f'variaciones_dcp_{timestamp}.xlsx'
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )


@bp.route('/stats/<int:product_id>', methods=['GET'])
def get_product_stats(product_id: int):
    """Get statistics for a product."""
    fecha_desde = request.args.get('fecha_desde', type=str)
    fecha_hasta = request.args.get('fecha_hasta', type=str)
    
    # Convert string dates to date objects if provided
    if fecha_desde:
        fecha_desde = date.fromisoformat(fecha_desde)
    if fecha_hasta:
        fecha_hasta = date.fromisoformat(fecha_hasta)
    
    # Convertir id sintético a id_variable e id_pais
    id_variable = product_id // 10000
    id_pais = product_id % 10000
    
    # Verificar que existe en maestro
    query_check = "SELECT id_variable, id_pais FROM maestro WHERE id_variable = ? AND id_pais = ?"
    fks_result = execute_query_single(query_check, (id_variable, id_pais))
    
    if not fks_result:
        return jsonify({'error': 'Product not found'}), 404
    
    # Build WHERE clause
    where_clause = "WHERE id_variable = ? AND id_pais = ?"
    params = [id_variable, id_pais]
    
    if fecha_desde:
        where_clause += " AND fecha >= ?"
        params.append(fecha_desde)
    if fecha_hasta:
        where_clause += " AND fecha <= ?"
        params.append(fecha_hasta)
    
    # Get min, max, and latest price
    query = f"""
        SELECT 
            MIN(valor) as precio_minimo,
            MAX(valor) as precio_maximo,
            MIN(fecha) as fecha_minima,
            MAX(fecha) as fecha_maxima,
            (SELECT valor FROM maestro_precios 
             WHERE id_variable = ? AND id_pais = ? {f"AND fecha <= ?" if fecha_hasta else ""} 
             ORDER BY fecha DESC LIMIT 1) as precio_actual
        FROM maestro_precios
        {where_clause}
    """
    
    if fecha_hasta:
        params_for_actual = [id_variable, id_pais, fecha_hasta]
    else:
        params_for_actual = [id_variable, id_pais]
    
    # Execute query with all params
    result = execute_query_single(query, tuple(params + params_for_actual))
    
    if not result:
        abort(404, description="Product not found or no data")
    
    # Calculate variation if we have date range
    variacion_periodo = None
    if fecha_desde and fecha_hasta and result.get('precio_minimo') is not None:
        # Get first and last price in range
        first_query = """
            SELECT valor FROM maestro_precios
            WHERE id_variable = ? AND id_pais = ? AND fecha >= ? AND fecha <= ?
            ORDER BY fecha ASC LIMIT 1
        """
        last_query = """
            SELECT valor FROM maestro_precios
            WHERE id_variable = ? AND id_pais = ? AND fecha >= ? AND fecha <= ?
            ORDER BY fecha DESC LIMIT 1
        """
        first_result = execute_query_single(first_query, (id_variable, id_pais, fecha_desde, fecha_hasta))
        last_result = execute_query_single(last_query, (id_variable, id_pais, fecha_desde, fecha_hasta))
        
        if first_result and last_result and first_result.get('valor') and last_result.get('valor'):
            first_price = first_result['valor']
            last_price = last_result['valor']
            if first_price > 0:
                variacion_periodo = ((last_price - first_price) / first_price) * 100.0
    
    return jsonify({
        'current_price': result.get('precio_actual'),
        'variation_last_month': variacion_periodo,
        'min_price_period': result.get('precio_minimo'),
        'max_price_period': result.get('precio_maximo')
    })


@bp.route('/products/prices/export', methods=['GET'])
def export_prices_to_excel():
    """Export prices for multiple products to Excel."""
    # Get product_ids from query params (can be multiple with same key)
    product_ids = request.args.getlist('product_ids[]', type=int)
    
    if not product_ids:
        abort(400, description="At least one product_id is required")
    
    fecha_desde = request.args.get('fecha_desde', type=str)
    fecha_hasta = request.args.get('fecha_hasta', type=str)
    
    # Convert string dates to date objects if provided
    if fecha_desde:
        fecha_desde = date.fromisoformat(fecha_desde)
    if fecha_hasta:
        fecha_hasta = date.fromisoformat(fecha_hasta)

    # Convertir product_ids sintéticos a (id_variable, id_pais)
    # id sintético = id_variable * 10000 + id_pais
    fks_map = {}
    for product_id in product_ids:
        id_variable = product_id // 10000
        id_pais = product_id % 10000
        # Verificar que existe en maestro
        query_check = "SELECT id_variable, id_pais FROM maestro WHERE id_variable = ? AND id_pais = ?"
        check_result = execute_query_single(query_check, (id_variable, id_pais))
        if check_result:
            fks_map[product_id] = (id_variable, id_pais)
    
    if not fks_map:
        abort(400, description="No products found with valid FKs")
    
    # Construir condiciones WHERE para FKs
    fks_conditions = []
    fks_params = []
    for id_var, id_pais in fks_map.values():
        fks_conditions.append("(mp.id_variable = ? AND mp.id_pais = ?)")
        fks_params.extend([id_var, id_pais])
    
    fks_where = " OR ".join(fks_conditions)
    
    # Build query with FKs
    if fecha_desde and fecha_hasta:
        query = f"""
            SELECT mp.id_variable, mp.id_pais, mp.fecha, mp.valor,
                   (m.id_variable * 10000 + m.id_pais) as id,
                   v.id_nombre_variable as nombre,
                   m.fuente, m.periodicidad
            FROM maestro_precios mp
            JOIN maestro m ON mp.id_variable = m.id_variable AND mp.id_pais = m.id_pais
            LEFT JOIN variables v ON m.id_variable = v.id_variable
            WHERE ({fks_where}) 
              AND mp.fecha BETWEEN ? AND ?
            ORDER BY (m.id_variable * 10000 + m.id_pais), mp.fecha ASC
        """
        params = tuple(fks_params) + (fecha_desde, fecha_hasta)
    elif fecha_desde:
        query = f"""
            SELECT mp.id_variable, mp.id_pais, mp.fecha, mp.valor,
                   (m.id_variable * 10000 + m.id_pais) as id,
                   v.id_nombre_variable as nombre,
                   m.fuente, m.periodicidad
            FROM maestro_precios mp
            JOIN maestro m ON mp.id_variable = m.id_variable AND mp.id_pais = m.id_pais
            LEFT JOIN variables v ON m.id_variable = v.id_variable
            WHERE ({fks_where}) 
              AND mp.fecha >= ?
            ORDER BY (m.id_variable * 10000 + m.id_pais), mp.fecha ASC
        """
        params = tuple(fks_params) + (fecha_desde,)
    elif fecha_hasta:
        query = f"""
            SELECT mp.id_variable, mp.id_pais, mp.fecha, mp.valor,
                   (m.id_variable * 10000 + m.id_pais) as id,
                   v.id_nombre_variable as nombre,
                   m.fuente, m.periodicidad
            FROM maestro_precios mp
            JOIN maestro m ON mp.id_variable = m.id_variable AND mp.id_pais = m.id_pais
            LEFT JOIN variables v ON m.id_variable = v.id_variable
            WHERE ({fks_where}) 
              AND mp.fecha <= ?
            ORDER BY (m.id_variable * 10000 + m.id_pais), mp.fecha ASC
        """
        params = tuple(fks_params) + (fecha_hasta,)
    else:
        query = f"""
            SELECT mp.id_variable, mp.id_pais, mp.fecha, mp.valor,
                   (m.id_variable * 10000 + m.id_pais) as id,
                   v.id_nombre_variable as nombre,
                   m.fuente, m.periodicidad
            FROM maestro_precios mp
            JOIN maestro m ON mp.id_variable = m.id_variable AND mp.id_pais = m.id_pais
            LEFT JOIN variables v ON m.id_variable = v.id_variable
            WHERE ({fks_where})
            ORDER BY (m.id_variable * 10000 + m.id_pais), mp.fecha ASC
        """
        params = tuple(fks_params)

    results = execute_query(query, params)
    
    # Group by product
    products_dict = {}
    for row in results:
        product_id = row['id']  # Synthetic ID: id_variable * 10000 + id_pais
        if product_id not in products_dict:
            products_dict[product_id] = {
                'product_id': row['id'],
                'product_name': row['nombre'],
                'unit': None,  # unidad no existe en nuevo schema
                'data': []
            }
        products_dict[product_id]['data'].append({
            'fecha': row['fecha'],
            'valor': row['valor']
        })
    
    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Precios Corrientes"
    
    # Header style
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Data style
    data_alignment = Alignment(horizontal="right", vertical="center")
    
    # Write headers
    row = 1
    ws['A1'] = 'Producto'
    ws['B1'] = 'Unidad'
    ws['C1'] = 'Fecha'
    ws['D1'] = 'Precio'
    
    # Apply header styles
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    # Write data
    row = 2
    for product_id, product_data in products_dict.items():
        product_name = product_data['product_name']
        unit = product_data['unit'] or ''
        
        for price_data in product_data['data']:
            ws[f'A{row}'] = product_name
            ws[f'B{row}'] = unit
            ws[f'C{row}'] = price_data['fecha']
            ws[f'D{row}'] = price_data['valor']
            
            # Apply data alignment to price column
            ws[f'D{row}'].alignment = data_alignment
            
            row += 1
    
    # Auto-adjust column widths
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    
    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Generate filename
    fecha_str = ""
    if fecha_desde and fecha_hasta:
        fecha_str = f"_{fecha_desde}_{fecha_hasta}"
    elif fecha_desde:
        fecha_str = f"_desde_{fecha_desde}"
    elif fecha_hasta:
        fecha_str = f"_hasta_{fecha_hasta}"
    
    filename = f"precios_corrientes{fecha_str}.xlsx"
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )
