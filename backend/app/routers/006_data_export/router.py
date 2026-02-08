"""API routes for data export."""
from datetime import date, datetime
from typing import List, Dict, Optional
from io import BytesIO
from flask import Blueprint, request, jsonify, send_file
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
import pandas as pd
from ...database import execute_query, execute_query_single

bp = Blueprint('data_export', __name__)


@bp.route('/export/families', methods=['GET'])
def get_families():
    """Obtiene todas las familias."""
    try:
        query = """
            SELECT id_familia, nombre_familia
            FROM familia
            ORDER BY nombre_familia
        """
        results = execute_query(query)
        # Asegurarse de que siempre devolvemos un array
        if not isinstance(results, list):
            results = []
        return jsonify(results)
    except Exception as e:
        # En caso de error, devolver array vacío
        print(f"Error en get_families: {e}")
        return jsonify([])


@bp.route('/export/subfamilies', methods=['GET'])
def get_subfamilies():
    """Obtiene subfamilias, opcionalmente filtradas por familia."""
    try:
        familia_id = request.args.get('familia_id', type=int)
        
        if familia_id:
            query = """
                SELECT sf.id_sub_familia, sf.nombre_sub_familia, sf.id_familia, f.nombre_familia
                FROM sub_familia sf
                LEFT JOIN familia f ON sf.id_familia = f.id_familia
                WHERE sf.id_familia = ?
                ORDER BY sf.nombre_sub_familia
            """
            results = execute_query(query, (familia_id,))
        else:
            query = """
                SELECT sf.id_sub_familia, sf.nombre_sub_familia, sf.id_familia, f.nombre_familia
                FROM sub_familia sf
                LEFT JOIN familia f ON sf.id_familia = f.id_familia
                ORDER BY f.nombre_familia, sf.nombre_sub_familia
            """
            results = execute_query(query)
        
        # Asegurarse de que siempre devolvemos un array
        if not isinstance(results, list):
            results = []
        
        return jsonify(results)
    except Exception as e:
        # En caso de error, devolver array vacío
        print(f"Error en get_subfamilies: {e}")
        return jsonify([])


@bp.route('/export/variables', methods=['GET'])
def get_variables():
    """Obtiene variables, opcionalmente filtradas por subfamilias."""
    try:
        subfamilia_ids = request.args.getlist('subfamilia_ids[]', type=int)
        
        if subfamilia_ids:
            # Filtrar por subfamilias seleccionadas
            placeholders = ','.join(['?'] * len(subfamilia_ids))
            query = f"""
                SELECT DISTINCT
                    v.id_variable,
                    v.id_nombre_variable as nombre,
                    sf.id_sub_familia,
                    sf.nombre_sub_familia,
                    f.id_familia,
                    f.nombre_familia
                FROM variables v
                LEFT JOIN sub_familia sf ON v.id_sub_familia = sf.id_sub_familia
                LEFT JOIN familia f ON sf.id_familia = f.id_familia
                WHERE v.id_sub_familia IN ({placeholders})
                ORDER BY f.nombre_familia, sf.nombre_sub_familia, v.id_nombre_variable
            """
            results = execute_query(query, tuple(subfamilia_ids))
        else:
            # Devolver todas las variables
            query = """
                SELECT DISTINCT
                    v.id_variable,
                    v.id_nombre_variable as nombre,
                    sf.id_sub_familia,
                    sf.nombre_sub_familia,
                    f.id_familia,
                    f.nombre_familia
                FROM variables v
                LEFT JOIN sub_familia sf ON v.id_sub_familia = sf.id_sub_familia
                LEFT JOIN familia f ON sf.id_familia = f.id_familia
                ORDER BY f.nombre_familia, sf.nombre_sub_familia, v.id_nombre_variable
            """
            results = execute_query(query)
        
        # Asegurarse de que siempre devolvemos un array
        if not isinstance(results, list):
            results = []
        
        return jsonify(results)
    except Exception as e:
        # En caso de error, devolver array vacío en lugar de objeto de error
        print(f"Error en get_variables: {e}")
        return jsonify([])


@bp.route('/export/countries', methods=['GET'])
def get_countries():
    """Obtiene países disponibles para las variables seleccionadas."""
    try:
        variable_ids = request.args.getlist('variable_ids[]', type=int)
        
        if not variable_ids:
            return jsonify([])
        
        placeholders = ','.join(['?'] * len(variable_ids))
        query = f"""
            SELECT DISTINCT
                pg.id_pais,
                pg.nombre_pais_grupo as nombre_pais
            FROM maestro_precios mp
            LEFT JOIN pais_grupo pg ON mp.id_pais = pg.id_pais
            WHERE mp.id_variable IN ({placeholders})
            AND pg.nombre_pais_grupo IS NOT NULL
            ORDER BY pg.nombre_pais_grupo
        """
        results = execute_query(query, tuple(variable_ids))
        
        # Asegurarse de que siempre devolvemos un array
        if not isinstance(results, list):
            results = []
        
        return jsonify(results)
    except Exception as e:
        # En caso de error, devolver array vacío
        print(f"Error en get_countries: {e}")
        return jsonify([])


@bp.route('/export/preview', methods=['GET'])
def get_preview():
    """Obtiene preview de datos en formato tidy (Variable, País, Fecha, Valor)."""
    try:
        variable_ids = request.args.getlist('variable_ids[]', type=int)
        pais_ids = request.args.getlist('pais_ids[]', type=int)
        fecha_desde = request.args.get('fecha_desde')
        fecha_hasta = request.args.get('fecha_hasta')
        
        if not variable_ids or not pais_ids:
            return jsonify({'error': 'Se requieren variables y países'}), 400
        
        # Construir condiciones
        var_placeholders = ','.join(['?'] * len(variable_ids))
        pais_placeholders = ','.join(['?'] * len(pais_ids))
        params = list(variable_ids) + list(pais_ids)
        
        where_clauses = [
            f"mp.id_variable IN ({var_placeholders})",
            f"mp.id_pais IN ({pais_placeholders})"
        ]
        
        if fecha_desde:
            where_clauses.append("mp.fecha >= ?")
            params.append(fecha_desde)
        
        if fecha_hasta:
            where_clauses.append("mp.fecha <= ?")
            params.append(fecha_hasta)
        
        # Query: obtener las últimas 10 filas
        query = f"""
            SELECT 
                v.id_nombre_variable as variable,
                pg.nombre_pais_grupo as pais,
                mp.fecha,
                mp.valor
            FROM maestro_precios mp
            LEFT JOIN variables v ON mp.id_variable = v.id_variable
            LEFT JOIN pais_grupo pg ON mp.id_pais = pg.id_pais
            WHERE {' AND '.join(where_clauses)}
            ORDER BY mp.fecha DESC, v.id_nombre_variable, pg.nombre_pais_grupo
            LIMIT 10
        """
        
        results = execute_query(query, tuple(params))
        
        if not results:
            return jsonify([])
        
        return jsonify(results)
    except Exception as e:
        # En caso de error, devolver array vacío
        print(f"Error en get_preview: {e}")
        import traceback
        traceback.print_exc()
        return jsonify([])


@bp.route('/export/download', methods=['GET'])
def download_excel():
    """Exporta datos a Excel en formato pivotado (fechas x países)."""
    try:
        variable_ids = request.args.getlist('variable_ids[]', type=int)
        pais_ids = request.args.getlist('pais_ids[]', type=int)
        fecha_desde = request.args.get('fecha_desde')
        fecha_hasta = request.args.get('fecha_hasta')
        
        if not variable_ids or not pais_ids:
            return jsonify({'error': 'Se requieren variables y países'}), 400
        
        # Construir query
        var_placeholders = ','.join(['?'] * len(variable_ids))
        pais_placeholders = ','.join(['?'] * len(pais_ids))
        params = list(variable_ids) + list(pais_ids)
        
        where_clauses = [
            f"mp.id_variable IN ({var_placeholders})",
            f"mp.id_pais IN ({pais_placeholders})"
        ]
        
        if fecha_desde:
            where_clauses.append("mp.fecha >= ?")
            params.append(fecha_desde)
        
        if fecha_hasta:
            where_clauses.append("mp.fecha <= ?")
            params.append(fecha_hasta)
        
        query = f"""
            SELECT 
                v.id_nombre_variable as variable,
                pg.nombre_pais_grupo as pais,
                mp.fecha,
                mp.valor
            FROM maestro_precios mp
            LEFT JOIN variables v ON mp.id_variable = v.id_variable
            LEFT JOIN pais_grupo pg ON mp.id_pais = pg.id_pais
            WHERE {' AND '.join(where_clauses)}
            ORDER BY v.id_nombre_variable, pg.nombre_pais_grupo, mp.fecha
        """
        
        results = execute_query(query, tuple(params))
        
        # Convertir a DataFrame para hacer pivot
        df = pd.DataFrame(results)
        
        if df.empty:
            return jsonify({'error': 'No hay datos para exportar'}), 400
        
        # Crear Excel con pandas
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            # Agrupar por variable
            for variable in df['variable'].unique():
                df_var = df[df['variable'] == variable]
                
                # Pivot: fechas en filas, países en columnas
                df_pivot = df_var.pivot_table(
                    index='fecha',
                    columns='pais',
                    values='valor',
                    aggfunc='first'  # Por si hay duplicados
                )
                
                # Resetear índice para que fecha sea columna
                df_pivot = df_pivot.reset_index()
                
                # Convertir fecha a formato dd-mm-yyyy
                df_pivot['fecha'] = pd.to_datetime(df_pivot['fecha']).dt.strftime('%d-%m-%Y')
                
                # Ordenar por fecha (convertir a datetime para ordenar, luego volver a string)
                df_pivot['fecha_dt'] = pd.to_datetime(df_pivot['fecha'], format='%d-%m-%Y')
                df_pivot = df_pivot.sort_values('fecha_dt')
                df_pivot = df_pivot.drop('fecha_dt', axis=1)
                
                # Renombrar columna fecha a "Fecha"
                df_pivot = df_pivot.rename(columns={'fecha': 'Fecha'})
                
                # Escribir a Excel
                sheet_name = variable[:31]  # Limitar nombre a 31 chars (límite de Excel)
                df_pivot.to_excel(writer, sheet_name=sheet_name, index=False)
                
                # Aplicar estilos a la hoja
                worksheet = writer.sheets[sheet_name]
                
                # Estilo para headers
                header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                header_font = Font(bold=True, color="FFFFFF")
                
                for cell in worksheet[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal="center")
                
                # Ajustar ancho de columnas
                worksheet.column_dimensions['A'].width = 12  # Fecha
                for col_idx, col in enumerate(df_pivot.columns[1:], start=2):
                    col_letter = worksheet.cell(row=1, column=col_idx).column_letter
                    worksheet.column_dimensions[col_letter].width = 15
        
        output.seek(0)
        filename = f"exportacion_datos_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500
