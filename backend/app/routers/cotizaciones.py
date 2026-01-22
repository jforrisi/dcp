"""API routes for LATAM exchange rates (cotizaciones)."""
from datetime import date, datetime, timedelta
from typing import List, Dict, Optional
from io import BytesIO
from flask import Blueprint, request, jsonify, abort, send_file
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from ..database import execute_query, execute_query_single

bp = Blueprint('cotizaciones', __name__)


def parse_fecha(fecha_val):
    """Convierte fecha de SQLite (puede ser date o datetime string) a date."""
    if isinstance(fecha_val, date):
        return fecha_val
    fecha_str = str(fecha_val)
    # Si tiene hora (contiene espacio), tomar solo la parte de fecha
    if ' ' in fecha_str:
        fecha_str = fecha_str.split(' ')[0]
    return date.fromisoformat(fecha_str)


@bp.route('/cotizaciones', methods=['GET'])
def get_cotizaciones():
    """
    Obtiene cotizaciones diarias de tipos de cambio LATAM.
    
    Query params:
    - product_ids[]: Lista de IDs de productos (cotizaciones)
    - fecha_desde: Fecha inicial (YYYY-MM-DD)
    - fecha_hasta: Fecha final (YYYY-MM-DD)
    """
    try:
        # Obtener parámetros
        product_ids = request.args.getlist('product_ids[]', type=int)
        fecha_desde_str = request.args.get('fecha_desde')
        fecha_hasta_str = request.args.get('fecha_hasta')
        
        if not product_ids:
            return jsonify({'error': 'Se requiere al menos un product_id'}), 400
        
        if not fecha_desde_str or not fecha_hasta_str:
            return jsonify({'error': 'Se requieren fecha_desde y fecha_hasta'}), 400
        
        fecha_desde = date.fromisoformat(fecha_desde_str)
        fecha_hasta = date.fromisoformat(fecha_hasta_str)
        
        if fecha_desde > fecha_hasta:
            return jsonify({'error': 'fecha_desde debe ser anterior a fecha_hasta'}), 400
        
        result = []
        
        # Obtener información de cada cotización
        placeholders = ','.join(['?'] * len(product_ids))
        query_products = f"""
            SELECT id, nombre, fuente, unidad, categoria
            FROM maestro
            WHERE id IN ({placeholders}) AND (activo = 1 OR CAST(activo AS INTEGER) = 1)
        """
        products = execute_query(query_products, tuple(product_ids))
        
        if not products:
            return jsonify({'error': 'No se encontraron cotizaciones activas'}), 404
        
        # Obtener datos de cada cotización
        for product in products:
            product_id = product['id']
            product_name = product['nombre']
            product_source = product.get('fuente', '')
            product_unidad = product.get('unidad', '')
            
            # Obtener precios diarios en el rango
            # Usar DATE() para normalizar fechas y asegurar comparación correcta
            # Esto maneja casos donde fecha puede tener componente de hora
            fecha_desde_str = fecha_desde.isoformat()
            fecha_hasta_str = fecha_hasta.isoformat()
            query_prices = """
                SELECT fecha, valor
                FROM maestro_precios
                WHERE maestro_id = ? 
                AND DATE(fecha) >= DATE(?)
                AND DATE(fecha) <= DATE(?)
                ORDER BY fecha ASC
            """
            prices = execute_query(query_prices, (product_id, fecha_desde_str, fecha_hasta_str))
            
            if not prices:
                continue
            
            # Normalizar fechas y preparar datos
            data = []
            for price_item in prices:
                fecha_obj = parse_fecha(price_item['fecha'])
                data.append({
                    'fecha': fecha_obj.isoformat(),
                    'valor': float(price_item['valor'])
                })
            
            # Calcular resumen
            if data:
                precio_inicial = data[0]['valor']
                precio_final = data[-1]['valor']
                variacion = 0.0
                if precio_inicial > 0:
                    variacion = ((precio_final - precio_inicial) / precio_inicial) * 100
                
                fecha_inicial = data[0]['fecha']
                fecha_final = data[-1]['fecha']
            else:
                precio_inicial = None
                precio_final = None
                variacion = 0.0
                fecha_inicial = None
                fecha_final = None
            
            result.append({
                'product_id': product_id,
                'product_name': product_name,
                'product_source': product_source,
                'unidad': product_unidad,
                'data': data,
                'summary': {
                    'precio_inicial': precio_inicial,
                    'precio_final': precio_final,
                    'variacion': variacion,
                    'fecha_inicial': fecha_inicial,
                    'fecha_final': fecha_final
                }
            })
        
        return jsonify(result)
    
    except ValueError as e:
        return jsonify({'error': f'Error en formato de fecha: {str(e)}'}), 400
    except Exception as e:
        return jsonify({'error': f'Error al obtener cotizaciones: {str(e)}'}), 500


@bp.route('/cotizaciones/products', methods=['GET'])
def get_cotizaciones_products():
    """
    Obtiene la lista de cotizaciones disponibles.
    Filtra usando solo la tabla maestro: tipo='M', periodicidad='D', es_cotizacion=1, activo=1
    """
    try:
        # Filtro único basado solo en la tabla maestro
        query = """
            SELECT id, nombre, fuente, unidad, categoria, periodicidad
            FROM maestro
            WHERE tipo = 'M' 
            AND periodicidad = 'D'
            AND es_cotizacion = 1
            AND (activo = 1 OR CAST(activo AS INTEGER) = 1)
            ORDER BY nombre
        """
        results = execute_query(query)
        return jsonify(results)
    
    except Exception as e:
        return jsonify({'error': f'Error al obtener productos: {str(e)}'}), 500


@bp.route('/cotizaciones/export', methods=['GET'])
def export_cotizaciones_to_excel():
    """
    Exporta cotizaciones a Excel.
    
    Query params:
    - product_ids[]: Lista de IDs de productos
    - fecha_desde: Fecha inicial (YYYY-MM-DD)
    - fecha_hasta: Fecha final (YYYY-MM-DD)
    """
    try:
        product_ids = request.args.getlist('product_ids[]', type=int)
        fecha_desde_str = request.args.get('fecha_desde')
        fecha_hasta_str = request.args.get('fecha_hasta')
        
        if not product_ids or not fecha_desde_str or not fecha_hasta_str:
            return jsonify({'error': 'Parámetros incompletos'}), 400
        
        fecha_desde = date.fromisoformat(fecha_desde_str)
        fecha_hasta = date.fromisoformat(fecha_hasta_str)
        
        # Crear workbook
        wb = Workbook()
        wb.remove(wb.active)  # Remover hoja por defecto
        
        # Hoja 1: Cotizaciones
        ws_cotizaciones = wb.create_sheet("Cotizaciones")
        
        # Encabezados
        headers = ['Fecha', 'País/Cotización', 'Valor', 'Unidad', 'Fuente']
        ws_cotizaciones.append(headers)
        
        # Estilo para encabezados
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for col in range(1, len(headers) + 1):
            cell = ws_cotizaciones.cell(row=1, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Obtener datos
        placeholders = ','.join(['?'] * len(product_ids))
        query_products = f"""
            SELECT id, nombre, fuente, unidad
            FROM maestro
            WHERE id IN ({placeholders}) AND activo = 1
        """
        products = execute_query(query_products, tuple(product_ids))
        
        row_num = 2
        for product in products:
            product_id = product['id']
            product_name = product['nombre']
            product_source = product.get('fuente', '')
            product_unidad = product.get('unidad', '')
            
            query_prices = """
                SELECT fecha, valor
                FROM maestro_precios
                WHERE maestro_id = ? AND fecha >= ? AND fecha <= ?
                ORDER BY fecha ASC
            """
            prices = execute_query(query_prices, (product_id, fecha_desde, fecha_hasta))
            
            for price_item in prices:
                fecha_obj = parse_fecha(price_item['fecha'])
                ws_cotizaciones.append([
                    fecha_obj.isoformat(),
                    product_name,
                    float(price_item['valor']),
                    product_unidad,
                    product_source
                ])
                row_num += 1
        
        # Ajustar ancho de columnas
        ws_cotizaciones.column_dimensions['A'].width = 12
        ws_cotizaciones.column_dimensions['B'].width = 30
        ws_cotizaciones.column_dimensions['C'].width = 15
        ws_cotizaciones.column_dimensions['D'].width = 15
        ws_cotizaciones.column_dimensions['E'].width = 20
        
        # Hoja 2: Metadatos
        ws_metadata = wb.create_sheet("Metadatos")
        ws_metadata.append(['Campo', 'Valor'])
        
        # Estilo para encabezados de metadatos
        for col in range(1, 3):
            cell = ws_metadata.cell(row=1, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        ws_metadata.append(['Fecha desde', fecha_desde_str])
        ws_metadata.append(['Fecha hasta', fecha_hasta_str])
        ws_metadata.append(['Total cotizaciones', len(products)])
        ws_metadata.append(['', ''])
        ws_metadata.append(['Cotización', 'ID'])
        
        for product in products:
            ws_metadata.append([product['nombre'], product['id']])
        
        ws_metadata.column_dimensions['A'].width = 25
        ws_metadata.column_dimensions['B'].width = 15
        
        # Guardar en BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Generar nombre de archivo
        filename = f"cotizaciones_latam_{fecha_desde_str}_{fecha_hasta_str}.xlsx"
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
    
    except ValueError as e:
        return jsonify({'error': f'Error en formato de fecha: {str(e)}'}), 400
    except Exception as e:
        return jsonify({'error': f'Error al exportar: {str(e)}'}), 500
