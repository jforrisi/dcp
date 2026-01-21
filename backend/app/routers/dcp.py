"""API routes for DCP (Dominant Currency Paradigm) index calculation."""
from datetime import date, datetime
from typing import List, Dict, Optional
from io import BytesIO
from flask import Blueprint, request, jsonify, abort, send_file
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from ..database import execute_query, execute_query_single

bp = Blueprint('dcp', __name__)

# IDs de variables macro según el sistema
TC_USD_ID = 6  # Tipo de cambio USD/UYU
TC_EUR_ID = 7  # Tipo de cambio EUR/UYU
IPC_ID = 11    # IPC general


def get_product_currency(product_id: int) -> Optional[str]:
    """
    Obtiene la moneda de un producto desde el maestro.
    
    Args:
        product_id: ID del producto en maestro
    
    Returns:
        Código de moneda en minúsculas ('usd', 'eur', 'uyu') o None si no se encuentra
    """
    query = "SELECT moneda FROM maestro WHERE id = ?"
    result = execute_query_single(query, (product_id,))
    if result and result.get('moneda'):
        return result['moneda'].lower()
    return None


def get_tc_for_product(product_id: int, tc_usd_monthly: Dict[date, float], 
                       tc_eur_monthly: Dict[date, float]) -> Optional[Dict[date, float]]:
    """
    Determina qué tipo de cambio usar para un producto basándose en su moneda.
    
    Args:
        product_id: ID del producto
        tc_usd_monthly: Diccionario de TC USD/UYU mensual
        tc_eur_monthly: Diccionario de TC EUR/UYU mensual
    
    Returns:
        Diccionario de tipo de cambio mensual apropiado, o None si no se puede determinar
    """
    moneda = get_product_currency(product_id)
    
    if moneda == 'eur':
        return tc_eur_monthly
    elif moneda == 'usd':
        return tc_usd_monthly
    elif moneda == 'uyu' or moneda is None:
        # Para UYU o si no hay moneda definida, retornar dict con valores 1.0
        # Esto es para productos que ya están en pesos uruguayos
        return None  # Se manejará como TC = 1.0 en el cálculo
    else:
        # Moneda desconocida, usar USD por defecto
        return tc_usd_monthly


def convert_to_monthly(series_data: List[Dict], periodicidad: str) -> List[Dict]:
    """
    Convierte una serie de datos a frecuencia mensual.
    
    Args:
        series_data: Lista de dicts con 'fecha' (date o string) y 'valor' (float)
        periodicidad: 'D', 'W', o 'M'
    
    Returns:
        Lista de dicts con fechas al primer día del mes y valores promediados
    """
    def parse_fecha(fecha_val):
        """Convierte fecha de SQLite (puede ser date o datetime string) a date."""
        if isinstance(fecha_val, date):
            return fecha_val
        fecha_str = str(fecha_val)
        # Si tiene hora (contiene espacio), tomar solo la parte de fecha
        if ' ' in fecha_str:
            fecha_str = fecha_str.split(' ')[0]
        return date.fromisoformat(fecha_str)
    
    # Normalizar todas las fechas a objetos date
    normalized_data = []
    for item in series_data:
        fecha_val = item['fecha']
        fecha_obj = parse_fecha(fecha_val)
        
        normalized_data.append({
            'fecha': fecha_obj,
            'valor': float(item['valor'])
        })
    
    if periodicidad == 'M':
        # Ya es mensual, pero puede haber múltiples valores por mes
        # Agrupar por año-mes y calcular promedio
        monthly_dict = {}
        
        for item in normalized_data:
            fecha_obj = item['fecha']
            year_month = (fecha_obj.year, fecha_obj.month)
            
            if year_month not in monthly_dict:
                monthly_dict[year_month] = {'sum': 0.0, 'count': 0}
            
            monthly_dict[year_month]['sum'] += item['valor']
            monthly_dict[year_month]['count'] += 1
        
        # Calcular promedios y crear lista de resultados
        result = []
        for (year, month), stats in sorted(monthly_dict.items()):
            if stats['count'] > 0:
                result.append({
                    'fecha': date(year, month, 1),
                    'valor': stats['sum'] / stats['count']
                })
        
        return result
    
    # Para D o W, agrupar por año-mes y calcular promedio
    monthly_dict = {}
    
    for item in normalized_data:
        fecha_obj = item['fecha']
        year_month = (fecha_obj.year, fecha_obj.month)
        
        if year_month not in monthly_dict:
            monthly_dict[year_month] = {'sum': 0.0, 'count': 0}
        
        monthly_dict[year_month]['sum'] += item['valor']
        monthly_dict[year_month]['count'] += 1
    
    # Calcular promedios y crear lista de resultados
    result = []
    for (year, month), stats in sorted(monthly_dict.items()):
        if stats['count'] > 0:
            result.append({
                'fecha': date(year, month, 1),
                'valor': stats['sum'] / stats['count']
            })
    
    return result


def get_macro_series(maestro_id: int, fecha_desde: date, fecha_hasta: date) -> Dict[date, float]:
    """
    Obtiene una serie macro (TC o IPC) y la convierte a mensual.
    
    Args:
        maestro_id: ID de la serie macro en maestro
        fecha_desde: Fecha inicial
        fecha_hasta: Fecha final
    
    Returns:
        Dict con fechas (primer día del mes) como keys y valores como values
    """
    # Obtener periodicidad de la serie macro
    query_maestro = "SELECT periodicidad FROM maestro WHERE id = ?"
    maestro_info = execute_query_single(query_maestro, (maestro_id,))
    
    if not maestro_info:
        return {}
    
    periodicidad = maestro_info['periodicidad']
    
    # Obtener datos de la serie FILTRADOS POR EL RANGO COMPLETO
    query = """
        SELECT fecha, valor
        FROM maestro_precios
        WHERE maestro_id = ? AND fecha >= ? AND fecha <= ?
        ORDER BY fecha ASC
    """
    try:
        raw_data = execute_query(query, (maestro_id, fecha_desde, fecha_hasta))
    except Exception as e:
        # Si hay error, retornar dict vacío
        return {}
    
    if not raw_data:
        return {}
    
    # Convertir a mensual (la función convert_to_monthly maneja la conversión de fechas)
    monthly_data = convert_to_monthly(raw_data, periodicidad)
    
    # Filtrar por rango usando comparación de año-mes
    fecha_desde_ym = (fecha_desde.year, fecha_desde.month)
    fecha_hasta_ym = (fecha_hasta.year, fecha_hasta.month)
    
    filtered_data = {}
    for item in monthly_data:
        mes_fecha = item['fecha']
        # Asegurar que sea date
        if not isinstance(mes_fecha, date):
            if isinstance(mes_fecha, str):
                mes_fecha = date.fromisoformat(mes_fecha)
            else:
                continue
        
        mes_fecha_ym = (mes_fecha.year, mes_fecha.month)
        if mes_fecha_ym >= fecha_desde_ym and mes_fecha_ym <= fecha_hasta_ym:
            filtered_data[mes_fecha] = item['valor']
    
    return filtered_data


@bp.route('/dcp/indices', methods=['GET'])
def get_dcp_indices():
    """
    Calcula índices DCP para productos seleccionados.
    
    Parámetros:
    - product_ids[]: Lista de IDs de productos
    - fecha_desde: Fecha inicial (YYYY-MM-DD)
    - fecha_hasta: Fecha final (YYYY-MM-DD)
    
    Retorna:
    JSON con índices normalizados a base 100
    """
    try:
        product_ids = request.args.getlist('product_ids[]', type=int)
        
        if not product_ids:
            return jsonify([])
        
        fecha_desde_str = request.args.get('fecha_desde', type=str)
        fecha_hasta_str = request.args.get('fecha_hasta', type=str)
        
        if not fecha_desde_str or not fecha_hasta_str:
            abort(400, description="fecha_desde and fecha_hasta are required")
        
        try:
            fecha_desde = date.fromisoformat(fecha_desde_str)
            fecha_hasta = date.fromisoformat(fecha_hasta_str)
        except ValueError as e:
            abort(400, description=f"Invalid date format: {str(e)}")
        
        # Obtener información de productos y servicios
        placeholders = ",".join("?" * len(product_ids))
        query_products = f"""
            SELECT id, nombre, periodicidad, fuente, moneda, nominal_real
            FROM maestro
            WHERE id IN ({placeholders}) AND tipo IN ('P', 'S', 'M') AND activo = 1
        """
        products = execute_query(query_products, tuple(product_ids))
        
        if not products:
            return jsonify([])
        
        # Obtener TC e IPC mensuales
        tc_usd_monthly = get_macro_series(TC_USD_ID, fecha_desde, fecha_hasta)
        tc_eur_monthly = get_macro_series(TC_EUR_ID, fecha_desde, fecha_hasta)
        ipc_monthly = get_macro_series(IPC_ID, fecha_desde, fecha_hasta)
        
        print(f"[DCP] Parámetros recibidos: product_ids={product_ids}, fecha_desde={fecha_desde}, fecha_hasta={fecha_hasta}")
        print(f"[DCP] Productos encontrados: {len(products)}")
        print(f"[DCP] TC USD meses: {len(tc_usd_monthly)}, TC EUR meses: {len(tc_eur_monthly)}, IPC meses: {len(ipc_monthly)}")
        
        if not ipc_monthly:
            print(f"[DCP] ERROR: No hay datos de IPC")
            return jsonify({
                'error': 'IPC data not available for the selected date range',
                'message': 'No hay datos de IPC disponibles para el rango de fechas seleccionado'
            }), 400
        
        result = []
        
        # Procesar cada producto
        for product in products:
            product_id = product['id']
            product_name = product['nombre']
            periodicidad = product['periodicidad']
            product_source = product.get('fuente', '')
            moneda = product.get('moneda')
            nominal_real = (product.get('nominal_real') or 'n').lower()
            
            # Obtener precios del producto (hasta fecha_hasta inclusive para mostrar datos hasta donde hay disponibles)
            query_prices = """
                SELECT fecha, valor
                FROM maestro_precios
                WHERE maestro_id = ? AND fecha <= ?
                ORDER BY fecha ASC
            """
            try:
                raw_prices = execute_query(query_prices, (product_id, fecha_hasta))
            except Exception as e:
                print(f"Error obteniendo precios para producto {product_id}: {str(e)}")
                continue
            
            if not raw_prices:
                continue
            
            # Convertir precios a mensual (la función maneja la conversión de fechas)
            prices_monthly = convert_to_monthly(raw_prices, periodicidad)
            
            # Filtrar precios mensuales por rango ANTES de calcular índices
            # Usar comparación de año-mes para asegurar que se incluya el último mes del rango
            fecha_desde_ym = (fecha_desde.year, fecha_desde.month)
            fecha_hasta_ym = (fecha_hasta.year, fecha_hasta.month)
            
            # Asegurar que todas las fechas sean objetos date antes de filtrar
            prices_monthly_filtered = []
            for p in prices_monthly:
                mes_fecha = p['fecha']
                # Convertir a date si es necesario
                if not isinstance(mes_fecha, date):
                    if isinstance(mes_fecha, str):
                        mes_fecha = date.fromisoformat(mes_fecha)
                    else:
                        continue
                
                mes_fecha_ym = (mes_fecha.year, mes_fecha.month)
                if mes_fecha_ym >= fecha_desde_ym and mes_fecha_ym <= fecha_hasta_ym:
                    p['fecha'] = mes_fecha  # Actualizar la fecha en el diccionario
                    prices_monthly_filtered.append(p)
            
            print(f"[DCP] Producto {product_id} ({product_name}): {len(prices_monthly)} precios mensuales, {len(prices_monthly_filtered)} filtrados")
            
            # Determinar qué tipo de cambio usar basándose en la moneda del producto
            if moneda == 'eur':
                tc_monthly = tc_eur_monthly
            elif moneda == 'usd':
                tc_monthly = tc_usd_monthly
            elif moneda == 'uyu' or moneda is None:
                # Para productos en UYU, crear TC=1.0 para todas las fechas de precios mensuales
                # Usar fechas de precios filtrados para asegurar coincidencia
                tc_monthly = {p['fecha']: 1.0 for p in prices_monthly_filtered}
            else:
                # Moneda desconocida, usar USD por defecto
                tc_monthly = tc_usd_monthly
            
            if not tc_monthly:
                print(f"[DCP] WARNING: No hay TC disponible para producto {product_id} ({product_name})")
                continue
            
            # Calcular índices con nominal/real:
            # base = precio × TC (según moneda)
            # si nominal_real == 'n' -> dividir por IPC; si 'r' -> no dividir por IPC
            indices_original = []
            for price_item in prices_monthly_filtered:
                mes_fecha = price_item['fecha']
                precio = float(price_item['valor'])
                
                # Asegurar que mes_fecha sea un objeto date para la comparación
                if not isinstance(mes_fecha, date):
                    if isinstance(mes_fecha, str):
                        mes_fecha = date.fromisoformat(mes_fecha)
                    else:
                        continue
                
                # Verificar que exista TC (y si es nominal, IPC) para este mes
                if mes_fecha in tc_monthly:
                    tc_valor = float(tc_monthly[mes_fecha])
                    base_valor = precio * tc_valor
                    
                    if nominal_real == 'r':
                        indices_original.append({'fecha': mes_fecha, 'valor': base_valor})
                    else:
                        if mes_fecha in ipc_monthly:
                            ipc_valor = float(ipc_monthly[mes_fecha])
                            if ipc_valor > 0:  # Evitar división por cero
                                indices_original.append({'fecha': mes_fecha, 'valor': base_valor / ipc_valor})
            
            if not indices_original:
                print(f"[DCP] WARNING: No se calcularon índices para producto {product_id} ({product_name})")
                continue
            
            print(f"[DCP] Producto {product_id} ({product_name}): {len(indices_original)} índices calculados")
            
            # Normalizar a base 100
            # Los índices ya están filtrados por rango, así que usamos todos los índices calculados
            # Usar comparación de año-mes para asegurar que se incluya el último mes del rango
            fecha_hasta_ym = (fecha_hasta.year, fecha_hasta.month)
            indices_filtered = [
                idx for idx in indices_original 
                if idx['fecha'] >= fecha_desde 
                and (idx['fecha'].year, idx['fecha'].month) <= fecha_hasta_ym
            ]
            
            if not indices_filtered:
                continue
            
            first_value = indices_filtered[0]['valor']
            
            if first_value == 0 or first_value is None:
                continue  # No se puede normalizar
            
            factor = 100.0 / first_value
            
            # Aplicar normalización - incluir todos los índices filtrados (hasta donde hay datos)
            indices_normalized = [
                {
                    'fecha': idx['fecha'].isoformat(),
                    'valor': idx['valor'] * factor
                }
                for idx in indices_filtered
            ]
            
            # Calcular información para la tabla de resumen
            # Precio inicial y final en moneda original
            precio_inicial = None
            precio_final = None
            fecha_inicial_producto = None
            fecha_final_producto = None
            variacion_precio_nominal = 0.0
            
            if prices_monthly_filtered and len(prices_monthly_filtered) > 0:
                try:
                    precio_inicial = float(prices_monthly_filtered[0]['valor'])
                    precio_final = float(prices_monthly_filtered[-1]['valor'])
                    fecha_inicial_producto = prices_monthly_filtered[0]['fecha']
                    fecha_final_producto = prices_monthly_filtered[-1]['fecha']
                    
                    # Calcular variación del precio nominal
                    if precio_inicial > 0:
                        variacion_precio_nominal = ((precio_final - precio_inicial) / precio_inicial) * 100
                    
                    print(f"[DCP] Producto {product_id} ({product_name}): precio_inicial={precio_inicial}, precio_final={precio_final}")
                    print(f"[DCP] Variación precio nominal: {variacion_precio_nominal}%")
                    print(f"[DCP] Fechas producto: {fecha_inicial_producto} a {fecha_final_producto}")
                except (ValueError, KeyError, IndexError) as e:
                    print(f"[DCP] ERROR calculando precios para producto {product_id}: {str(e)}")
                    precio_inicial = None
                    precio_final = None
            else:
                print(f"[DCP] WARNING: No hay precios filtrados para producto {product_id} ({product_name})")
            
            # Filtrar series macro por el rango del producto ANTES de calcular variaciones
            # Usar las fechas reales del producto (no el rango completo del filtro)
            variacion_tc = 0.0
            variacion_ipc = 0.0
            
            if fecha_inicial_producto and fecha_final_producto:
                # Rango real del producto (por año-mes), dentro del rango que pidió el usuario
                # Ej: si usuario pide hasta dic-2025 pero el producto llega a set-2025, end_ym = (2025, 9)
                start_ym = (fecha_inicial_producto.year, fecha_inicial_producto.month)
                end_ym = (fecha_final_producto.year, fecha_final_producto.month)

                # Filtrar TC según la moneda del producto
                # Nota: las series macro suelen estar en fechas date(year, month, 1).
                # Para ser robustos, calculamos variación por año-mes, tomando el primer y último mes disponibles.
                if moneda == 'uyu' or moneda is None:
                    variacion_tc = 0.0
                else:
                    tc_series = None
                    if moneda == 'eur':
                        tc_series = tc_eur_monthly
                    elif moneda == 'usd':
                        tc_series = tc_usd_monthly

                    if tc_series:
                        tc_by_ym = {(f.year, f.month): float(v) for f, v in tc_series.items() if isinstance(f, date)}
                        tc_ym_in_range = sorted([ym for ym in tc_by_ym.keys() if start_ym <= ym <= end_ym])
                        if len(tc_ym_in_range) >= 2:
                            tc_inicial = tc_by_ym[tc_ym_in_range[0]]
                            tc_final = tc_by_ym[tc_ym_in_range[-1]]
                            if tc_inicial > 0:
                                variacion_tc = ((tc_final / tc_inicial) - 1.0) * 100
                            print(f"[DCP] Variación TC: {variacion_tc}% (de {tc_ym_in_range[0]} a {tc_ym_in_range[-1]})")
                        else:
                            variacion_tc = 0.0
                    else:
                        variacion_tc = 0.0
                
                # Variación de IPC en el período (usando fechas filtradas del producto)
                nominal_real_norm = (nominal_real or 'n')
                if isinstance(nominal_real_norm, str):
                    nominal_real_norm = nominal_real_norm.strip().lower()
                else:
                    nominal_real_norm = 'n'

                if nominal_real_norm != 'n':
                    # Variable real => no aplica IPC
                    variacion_ipc = 0.0
                else:
                    # Variable nominal => usar (IPC_final / IPC_inicial - 1) * 100
                    if ipc_monthly:
                        ipc_by_ym = {(f.year, f.month): float(v) for f, v in ipc_monthly.items() if isinstance(f, date)}
                        ipc_ym_in_range = sorted([ym for ym in ipc_by_ym.keys() if start_ym <= ym <= end_ym])
                        if len(ipc_ym_in_range) >= 2:
                            ipc_inicial = ipc_by_ym[ipc_ym_in_range[0]]
                            ipc_final = ipc_by_ym[ipc_ym_in_range[-1]]
                            if ipc_inicial > 0:
                                variacion_ipc = ((ipc_final / ipc_inicial) - 1.0) * 100
                            print(f"[DCP] Variación IPC: {variacion_ipc}% (de {ipc_ym_in_range[0]} a {ipc_ym_in_range[-1]})")
                        else:
                            variacion_ipc = 0.0
                    else:
                        variacion_ipc = 0.0
            
            # Variación real del índice DCP
            variacion_real = 0.0
            if indices_normalized and len(indices_normalized) >= 2:
                indice_inicial = indices_normalized[0]['valor']
                indice_final = indices_normalized[-1]['valor']
                if indice_inicial > 0:
                    variacion_real = ((indice_final - indice_inicial) / indice_inicial) * 100
            
            # Validar fórmula: (1+var_precio_nominal) * (1+var_tc) / (1+var_inflacion) = (1+var_real)
            if precio_inicial and precio_final:
                var_precio_decimal = variacion_precio_nominal / 100.0
                var_tc_decimal = variacion_tc / 100.0
                var_inflacion_decimal = variacion_ipc / 100.0
                var_real_decimal = variacion_real / 100.0
                
                # Evitar división por cero si var_inflacion es -100%
                if var_inflacion_decimal != -1.0:
                    lado_izquierdo = (1 + var_precio_decimal) * (1 + var_tc_decimal) / (1 + var_inflacion_decimal)
                    lado_derecho = 1 + var_real_decimal
                    
                    diferencia = abs(lado_izquierdo - lado_derecho)
                    if diferencia > 0.01:  # Tolerancia de 0.01 (1%)
                        print(f"[DCP] WARNING: Fórmula no se cumple para producto {product_id} ({product_name})")
                        print(f"[DCP]   (1+var_precio)* (1+var_tc)/(1+var_inflacion) = {lado_izquierdo:.6f}")
                        print(f"[DCP]   (1+var_real) = {lado_derecho:.6f}")
                        print(f"[DCP]   Diferencia: {diferencia:.6f}")
                    else:
                        print(f"[DCP] ✓ Fórmula validada para producto {product_id} ({product_name})")
            
            result.append({
                'product_id': product_id,
                'product_name': product_name,
                'product_source': product_source,
                'moneda': moneda or 'uyu',
                'nominal_real': nominal_real,
                'data': indices_normalized,
                'summary': {
                    'precio_inicial': precio_inicial,
                    'precio_final': precio_final,
                    'variacion_precio_nominal': variacion_precio_nominal,
                    'variacion_tc': variacion_tc,
                    'variacion_ipc': variacion_ipc,
                    'variacion_real': variacion_real,
                    'fecha_inicial': fecha_inicial_producto.isoformat() if fecha_inicial_producto else None,
                    'fecha_final': fecha_final_producto.isoformat() if fecha_final_producto else None
                }
            })
        
        print(f"[DCP] Total productos procesados exitosamente: {len(result)}")
        return jsonify(result)
    except Exception as e:
        import traceback
        error_trace = traceback.format_exc()
        print(f"Error en get_dcp_indices: {str(e)}")
        print(error_trace)
        return jsonify({
            'error': 'Error calculating DCP indices',
            'message': str(e)
        }), 500


@bp.route('/dcp/indices/export', methods=['GET'])
def export_dcp_indices_to_excel():
    """
    Exporta índices DCP a Excel con 3 hojas.
    
    Parámetros: iguales a /dcp/indices
    Retorna: archivo Excel
    """
    product_ids = request.args.getlist('product_ids[]', type=int)
    
    if not product_ids:
        abort(400, description="At least one product_id is required")
    
    fecha_desde_str = request.args.get('fecha_desde', type=str)
    fecha_hasta_str = request.args.get('fecha_hasta', type=str)
    
    if not fecha_desde_str or not fecha_hasta_str:
        abort(400, description="fecha_desde and fecha_hasta are required")
    
    fecha_desde = date.fromisoformat(fecha_desde_str)
    fecha_hasta = date.fromisoformat(fecha_hasta_str)
    
    # Obtener datos (similar a /indices pero necesitamos originales también)
    placeholders = ",".join("?" * len(product_ids))
    query_products = f"""
        SELECT id, nombre, periodicidad, fuente, moneda, nominal_real
        FROM maestro
        WHERE id IN ({placeholders}) AND tipo IN ('P', 'S', 'M') AND activo = 1
    """
    products = execute_query(query_products, tuple(product_ids))
    
    if not products:
        abort(400, description="No valid products found")
    
    # Obtener TC e IPC mensuales
    tc_usd_monthly = get_macro_series(TC_USD_ID, fecha_desde, fecha_hasta)
    tc_eur_monthly = get_macro_series(TC_EUR_ID, fecha_desde, fecha_hasta)
    ipc_monthly = get_macro_series(IPC_ID, fecha_desde, fecha_hasta)
    
    if not ipc_monthly:
        abort(400, description="IPC data not available")
    
    # Calcular índices originales y normalizados para cada producto
    all_indices_original = {}  # {product_id: [(fecha, valor), ...]}
    all_indices_normalized = {}  # {product_id: [(fecha, valor), ...]}
    all_prices_original = {}  # {product_id: [(fecha, precio), ...]}
    product_names = {}  # {product_id: nombre}
    
    for product in products:
        product_id = product['id']
        product_name = product['nombre']
        periodicidad = product['periodicidad']
        product_source = product.get('fuente', '')
        moneda = product.get('moneda')
        nominal_real = (product.get('nominal_real') or 'n').lower()
        product_names[product_id] = product_name
        
        # Obtener precios (hasta fecha_hasta inclusive para mostrar datos hasta donde hay disponibles)
        query_prices = """
            SELECT fecha, valor
            FROM maestro_precios
            WHERE maestro_id = ? AND fecha <= ?
            ORDER BY fecha ASC
        """
        raw_prices = execute_query(query_prices, (product_id, fecha_hasta))
        
        if not raw_prices:
            continue
        
        # Normalizar fechas usando la misma lógica que convert_to_monthly
        for item in raw_prices:
            fecha_val = item['fecha']
            if isinstance(fecha_val, date):
                continue  # Ya es date
            fecha_str = str(fecha_val)
            # Si tiene hora (contiene espacio), tomar solo la parte de fecha
            if ' ' in fecha_str:
                fecha_str = fecha_str.split(' ')[0]
            item['fecha'] = date.fromisoformat(fecha_str)
        
        prices_monthly = convert_to_monthly(raw_prices, periodicidad)
        
        # Guardar precios mensuales SOLO del rango seleccionado (fecha_desde a fecha_hasta)
        # Usar comparación de año-mes para asegurar que se incluya el último mes del rango
        fecha_desde_ym = (fecha_desde.year, fecha_desde.month)
        fecha_hasta_ym = (fecha_hasta.year, fecha_hasta.month)
        
        for price_item in prices_monthly:
            mes_fecha = price_item['fecha']
            precio = float(price_item['valor'])
            
            if not isinstance(mes_fecha, date):
                if isinstance(mes_fecha, str):
                    mes_fecha = date.fromisoformat(mes_fecha)
                else:
                    continue
            
            # Filtrar solo precios dentro del rango seleccionado usando comparación de año-mes
            mes_fecha_ym = (mes_fecha.year, mes_fecha.month)
            if mes_fecha_ym >= fecha_desde_ym and mes_fecha_ym <= fecha_hasta_ym:
                if product_id not in all_prices_original:
                    all_prices_original[product_id] = []
                all_prices_original[product_id].append((mes_fecha, precio))
        
        # Determinar qué tipo de cambio usar basándose en la moneda del producto
        if moneda == 'eur':
            tc_monthly = tc_eur_monthly
        elif moneda == 'usd':
            tc_monthly = tc_usd_monthly
        elif moneda == 'uyu' or moneda is None:
            # Para productos en UYU, crear TC=1.0 para todas las fechas de precios mensuales
            # Filtrar precios por rango primero
            prices_filtered = []
            for p in prices_monthly:
                mes_fecha = p['fecha']
                # Asegurar que mes_fecha sea un objeto date
                if not isinstance(mes_fecha, date):
                    if isinstance(mes_fecha, str):
                        mes_fecha = date.fromisoformat(mes_fecha)
                    else:
                        continue
                mes_fecha_ym = (mes_fecha.year, mes_fecha.month)
                if mes_fecha_ym >= fecha_desde_ym and mes_fecha_ym <= fecha_hasta_ym:
                    prices_filtered.append({'fecha': mes_fecha, 'valor': p['valor']})
            tc_monthly = {p['fecha']: 1.0 for p in prices_filtered}
        else:
            # Moneda desconocida, usar USD por defecto
            tc_monthly = tc_usd_monthly
        
        if not tc_monthly:
            continue
        
        # Calcular índices originales con nominal/real:
        # base = precio × TC; si nominal_real == 'n' -> dividir por IPC; si 'r' -> no dividir
        indices_orig = []
        for price_item in prices_monthly:
            mes_fecha = price_item['fecha']
            precio = price_item['valor']
            
            if mes_fecha in tc_monthly:
                tc_valor = tc_monthly[mes_fecha]
                base_valor = precio * tc_valor
                if nominal_real == 'r':
                    indices_orig.append((mes_fecha, base_valor))
                else:
                    if mes_fecha in ipc_monthly:
                        ipc_valor = ipc_monthly[mes_fecha]
                        if ipc_valor > 0:
                            indices_orig.append((mes_fecha, base_valor / ipc_valor))
        
        if not indices_orig:
            continue
        
        # Filtrar índices originales por rango antes de guardar
        # Esto asegura que la hoja "Índices Originales" solo muestre el rango seleccionado
        indices_orig_filtered = [idx for idx in indices_orig if idx[0] >= fecha_desde and idx[0] <= fecha_hasta]
        all_indices_original[product_id] = indices_orig_filtered
        
        # Normalizar usando exactamente el mismo rango filtrado (consistente con "Índices Originales")
        indices_filtered = indices_orig_filtered
        if not indices_filtered:
            continue
        
        first_value = indices_filtered[0][1]
        if first_value == 0 or first_value is None:
            continue
        
        factor = 100.0 / first_value
        indices_norm = [(fecha, valor * factor) for fecha, valor in indices_filtered]
        all_indices_normalized[product_id] = indices_norm
    
    # Crear Excel
    wb = Workbook()
    
    # Estilos
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_alignment = Alignment(horizontal="center", vertical="center")
    data_alignment = Alignment(horizontal="right", vertical="center")
    
    # Hoja 1: Índices Normalizados
    ws1 = wb.active
    ws1.title = "Índices Normalizados"
    
    # Encabezados
    ws1['A1'] = 'Fecha'
    ws1['A1'].fill = header_fill
    ws1['A1'].font = header_font
    ws1['A1'].alignment = header_alignment
    
    col = 2
    for product_id in sorted(all_indices_normalized.keys()):
        cell = ws1.cell(row=1, column=col)
        cell.value = product_names[product_id]
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        col += 1
    
    # Obtener todas las fechas únicas
    all_dates = set()
    for indices in all_indices_normalized.values():
        all_dates.update([idx[0] for idx in indices])
    sorted_dates = sorted(all_dates)
    
    # Escribir datos
    row = 2
    for fecha in sorted_dates:
        ws1.cell(row=row, column=1).value = fecha
        col = 2
        for product_id in sorted(all_indices_normalized.keys()):
            indices = all_indices_normalized[product_id]
            valor = next((idx[1] for idx in indices if idx[0] == fecha), None)
            if valor is not None:
                cell = ws1.cell(row=row, column=col)
                cell.value = valor
                cell.alignment = data_alignment
            col += 1
        row += 1
    
    # Hoja 2: Índices Originales
    ws2 = wb.create_sheet("Índices Originales")
    
    ws2['A1'] = 'Fecha'
    ws2['A1'].fill = header_fill
    ws2['A1'].font = header_font
    ws2['A1'].alignment = header_alignment
    
    col = 2
    for product_id in sorted(all_indices_original.keys()):
        cell = ws2.cell(row=1, column=col)
        cell.value = product_names[product_id]
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        col += 1
    
    all_dates_orig = set()
    for indices in all_indices_original.values():
        all_dates_orig.update([idx[0] for idx in indices])
    sorted_dates_orig = sorted(all_dates_orig)
    
    row = 2
    for fecha in sorted_dates_orig:
        ws2.cell(row=row, column=1).value = fecha
        col = 2
        for product_id in sorted(all_indices_original.keys()):
            indices = all_indices_original[product_id]
            valor = next((idx[1] for idx in indices if idx[0] == fecha), None)
            if valor is not None:
                cell = ws2.cell(row=row, column=col)
                cell.value = valor
                cell.alignment = data_alignment
            col += 1
        row += 1
    
    # Hoja 3: Precios Originales (solo del rango seleccionado, con TC e IPC)
    ws3 = wb.create_sheet("Precios Originales")
    ws3['A1'] = 'Fecha'
    ws3['A1'].fill = header_fill
    ws3['A1'].font = header_font
    ws3['A1'].alignment = header_alignment
    
    # Encabezados de productos
    product_cols = {}
    col = 2
    for product_id in sorted(all_prices_original.keys()):
        cell = ws3.cell(row=1, column=col)
        cell.value = product_names[product_id]
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        product_cols[product_id] = col
        col += 1
    
    # Encabezados de IPC y TC
    ipc_col = col
    ws3.cell(row=1, column=ipc_col).value = 'IPC'
    ws3.cell(row=1, column=ipc_col).fill = header_fill
    ws3.cell(row=1, column=ipc_col).font = header_font
    ws3.cell(row=1, column=ipc_col).alignment = header_alignment
    col += 1
    
    tc_usd_col = col
    ws3.cell(row=1, column=tc_usd_col).value = 'TC USD/UYU'
    ws3.cell(row=1, column=tc_usd_col).fill = header_fill
    ws3.cell(row=1, column=tc_usd_col).font = header_font
    ws3.cell(row=1, column=tc_usd_col).alignment = header_alignment
    col += 1
    
    tc_eur_col = col
    ws3.cell(row=1, column=tc_eur_col).value = 'TC EUR/UYU'
    ws3.cell(row=1, column=tc_eur_col).fill = header_fill
    ws3.cell(row=1, column=tc_eur_col).font = header_font
    ws3.cell(row=1, column=tc_eur_col).alignment = header_alignment
    
    # Obtener todas las fechas del rango (solo las que están en el rango seleccionado)
    all_dates_prices = set()
    for prices in all_prices_original.values():
        all_dates_prices.update([p[0] for p in prices])
    # También incluir fechas de IPC y TC que estén en el rango
    all_dates_prices.update([f for f in ipc_monthly.keys() if f >= fecha_desde and f <= fecha_hasta])
    all_dates_prices.update([f for f in tc_usd_monthly.keys() if f >= fecha_desde and f <= fecha_hasta])
    all_dates_prices.update([f for f in tc_eur_monthly.keys() if f >= fecha_desde and f <= fecha_hasta])
    sorted_dates_prices = sorted(all_dates_prices)
    
    row = 2
    for fecha in sorted_dates_prices:
        ws3.cell(row=row, column=1).value = fecha
        
        # Escribir precios de productos
        for product_id in sorted(all_prices_original.keys()):
            prices = all_prices_original[product_id]
            valor = next((p[1] for p in prices if p[0] == fecha), None)
            if valor is not None:
                cell = ws3.cell(row=row, column=product_cols[product_id])
                cell.value = valor
                cell.number_format = '0.00'
                cell.alignment = data_alignment
        
        # Escribir IPC
        ipc_valor = ipc_monthly.get(fecha)
        if ipc_valor is not None:
            cell = ws3.cell(row=row, column=ipc_col)
            cell.value = ipc_valor
            cell.number_format = '0.00'
            cell.alignment = data_alignment
        
        # Escribir TC USD/UYU
        tc_usd_valor = tc_usd_monthly.get(fecha)
        if tc_usd_valor is not None:
            cell = ws3.cell(row=row, column=tc_usd_col)
            cell.value = tc_usd_valor
            cell.number_format = '0.00'
            cell.alignment = data_alignment
        
        # Escribir TC EUR/UYU
        tc_eur_valor = tc_eur_monthly.get(fecha)
        if tc_eur_valor is not None:
            cell = ws3.cell(row=row, column=tc_eur_col)
            cell.value = tc_eur_valor
            cell.number_format = '0.00'
            cell.alignment = data_alignment
        
        row += 1
    
    # Hoja 4: Metadatos
    ws4 = wb.create_sheet("Metadatos")
    
    ws4['A1'] = 'Campo'
    ws4['A1'].fill = header_fill
    ws4['A1'].font = header_font
    ws4['B1'] = 'Valor'
    ws4['B1'].fill = header_fill
    ws4['B1'].font = header_font
    
    metadata = [
        ('Fecha de exportación', datetime.now().strftime('%Y-%m-%d %H:%M:%S')),
        ('Rango de fechas', f'{fecha_desde} a {fecha_hasta}'),
        ('Productos incluidos', ', '.join([product_names[pid] for pid in sorted(all_indices_normalized.keys())])),
    ]
    
    row = 2
    for campo, valor in metadata:
        ws4.cell(row=row, column=1).value = campo
        ws4.cell(row=row, column=2).value = valor
        row += 1
    
    # Agregar fórmulas por producto
    ws4.cell(row=row, column=1).value = 'Fórmulas aplicadas'
    ws4.cell(row=row, column=1).font = Font(bold=True)
    row += 1
    
    for product_id in sorted(all_indices_normalized.keys()):
        moneda = get_product_currency(product_id)
        if moneda == 'eur':
            tc_type = 'EUR/UYU'
        elif moneda == 'usd':
            tc_type = 'USD/UYU'
        elif moneda == 'uyu' or moneda is None:
            tc_type = 'UYU (TC=1.0)'
        else:
            tc_type = 'USD/UYU'
        ws4.cell(row=row, column=1).value = product_names[product_id]
        ws4.cell(row=row, column=2).value = f'Precio internacional × TC {tc_type} / IPC'
        row += 1
    
    # Ajustar anchos de columna
    for ws in [ws1, ws2]:
        ws.column_dimensions['A'].width = 15
        for col in range(2, len(all_indices_normalized) + 2):
            ws.column_dimensions[chr(64 + col)].width = 25
    
    # Para ws3 (Precios Originales), ajustar considerando productos + IPC + TC USD + TC EUR
    ws3.column_dimensions['A'].width = 15
    num_product_cols = len(all_prices_original)
    for col in range(2, num_product_cols + 2):
        ws3.column_dimensions[chr(64 + col)].width = 25
    # IPC, TC USD/UYU, TC EUR/UYU
    for col in range(num_product_cols + 2, num_product_cols + 5):
        ws3.column_dimensions[chr(64 + col)].width = 15
    
    ws4.column_dimensions['A'].width = 30
    ws4.column_dimensions['B'].width = 50
    
    # Guardar a BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Nombre de archivo
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f'indices_dcp_{timestamp}.xlsx'
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )
