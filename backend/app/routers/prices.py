"""API routes for price data."""
from datetime import date
from typing import List, Optional
from io import BytesIO
from flask import Blueprint, request, jsonify, abort, send_file
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from ..database import execute_query, execute_query_single

bp = Blueprint('prices', __name__)


@bp.route('/products', methods=['GET'])
def get_products():
    """Get all active products (tipo='P')."""
    query = """
        SELECT id, nombre, tipo, unidad, categoria, fuente, periodicidad, activo
        FROM maestro
        WHERE tipo = 'P' AND activo = 1
        ORDER BY nombre
    """
    results = execute_query(query)
    return jsonify(results)


@bp.route('/products/<int:product_id>/prices', methods=['GET'])
def get_product_prices(product_id: int):
    """Get prices for a single product within a date range."""
    fecha_desde = request.args.get('fecha_desde', type=str)
    fecha_hasta = request.args.get('fecha_hasta', type=str)
    
    # Convert string dates to date objects if provided
    if fecha_desde:
        fecha_desde = date.fromisoformat(fecha_desde)
    if fecha_hasta:
        fecha_hasta = date.fromisoformat(fecha_hasta)
    
    if fecha_desde and fecha_hasta:
        query = """
            SELECT fecha, valor
            FROM maestro_precios
            WHERE maestro_id = ? AND fecha BETWEEN ? AND ?
            ORDER BY fecha ASC
        """
        params = (product_id, fecha_desde, fecha_hasta)
    elif fecha_desde:
        query = """
            SELECT fecha, valor
            FROM maestro_precios
            WHERE maestro_id = ? AND fecha >= ?
            ORDER BY fecha ASC
        """
        params = (product_id, fecha_desde)
    elif fecha_hasta:
        query = """
            SELECT fecha, valor
            FROM maestro_precios
            WHERE maestro_id = ? AND fecha <= ?
            ORDER BY fecha ASC
        """
        params = (product_id, fecha_hasta)
    else:
        query = """
            SELECT fecha, valor
            FROM maestro_precios
            WHERE maestro_id = ?
            ORDER BY fecha ASC
        """
        params = (product_id,)

    results = execute_query(query, params)
    return jsonify(results)


@bp.route('/products/prices', methods=['GET'])
def get_multiple_products_prices():
    """Get prices for multiple products within a date range."""
    # Get product_ids from query params (can be multiple with same key)
    product_ids = request.args.getlist('product_ids[]', type=int)
    
    if not product_ids:
        return jsonify([])
    
    fecha_desde = request.args.get('fecha_desde', type=str)
    fecha_hasta = request.args.get('fecha_hasta', type=str)
    
    # Convert string dates to date objects if provided
    if fecha_desde:
        fecha_desde = date.fromisoformat(fecha_desde)
    if fecha_hasta:
        fecha_hasta = date.fromisoformat(fecha_hasta)

    # Build query with IN clause
    placeholders = ",".join("?" * len(product_ids))
    
    if fecha_desde and fecha_hasta:
        query = f"""
            SELECT mp.maestro_id, mp.fecha, mp.valor,
                   m.id, m.nombre, m.tipo, m.unidad, m.categoria, m.fuente, m.periodicidad
            FROM maestro_precios mp
            JOIN maestro m ON mp.maestro_id = m.id
            WHERE mp.maestro_id IN ({placeholders}) 
              AND mp.fecha BETWEEN ? AND ?
            ORDER BY mp.maestro_id, mp.fecha ASC
        """
        params = tuple(product_ids) + (fecha_desde, fecha_hasta)
    elif fecha_desde:
        query = f"""
            SELECT mp.maestro_id, mp.fecha, mp.valor,
                   m.id, m.nombre, m.tipo, m.unidad, m.categoria, m.fuente, m.periodicidad
            FROM maestro_precios mp
            JOIN maestro m ON mp.maestro_id = m.id
            WHERE mp.maestro_id IN ({placeholders}) 
              AND mp.fecha >= ?
            ORDER BY mp.maestro_id, mp.fecha ASC
        """
        params = tuple(product_ids) + (fecha_desde,)
    elif fecha_hasta:
        query = f"""
            SELECT mp.maestro_id, mp.fecha, mp.valor,
                   m.id, m.nombre, m.tipo, m.unidad, m.categoria, m.fuente, m.periodicidad
            FROM maestro_precios mp
            JOIN maestro m ON mp.maestro_id = m.id
            WHERE mp.maestro_id IN ({placeholders}) 
              AND mp.fecha <= ?
            ORDER BY mp.maestro_id, mp.fecha ASC
        """
        params = tuple(product_ids) + (fecha_hasta,)
    else:
        query = f"""
            SELECT mp.maestro_id, mp.fecha, mp.valor,
                   m.id, m.nombre, m.tipo, m.unidad, m.categoria, m.fuente, m.periodicidad
            FROM maestro_precios mp
            JOIN maestro m ON mp.maestro_id = m.id
            WHERE mp.maestro_id IN ({placeholders})
            ORDER BY mp.maestro_id, mp.fecha ASC
        """
        params = tuple(product_ids)

    results = execute_query(query, params)
    
    # Group by product and format response
    products_dict = {}
    for row in results:
        product_id = row['maestro_id']
        if product_id not in products_dict:
            products_dict[product_id] = {
                'product_id': row['id'],
                'product_name': row['nombre'],
                'unit': row['unidad'],
                'data': []
            }
        products_dict[product_id]['data'].append({
            'fecha': row['fecha'],
            'valor': row['valor']
        })
    
    return jsonify(list(products_dict.values()))


@bp.route('/variations', methods=['GET'])
def get_price_variations():
    """Calculate price variations for all products in a date range."""
    fecha_desde = request.args.get('fecha_desde', type=str)
    fecha_hasta = request.args.get('fecha_hasta', type=str)
    order_by = request.args.get('order_by', default='desc', type=str)
    
    if not fecha_desde or not fecha_hasta:
        abort(400, description="fecha_desde and fecha_hasta are required")
    
    if order_by not in ['asc', 'desc']:
        order_by = 'desc'
    
    # Convert string dates to date objects
    fecha_desde = date.fromisoformat(fecha_desde)
    fecha_hasta = date.fromisoformat(fecha_hasta)
    
    query = """
        WITH precios_iniciales AS (
            SELECT 
                mp1.maestro_id, 
                mp1.valor as precio_inicial,
                mp1.fecha as fecha_inicial
            FROM maestro_precios mp1
            WHERE mp1.fecha = (
                SELECT MIN(mp2.fecha)
                FROM maestro_precios mp2
                WHERE mp2.maestro_id = mp1.maestro_id
                  AND mp2.fecha >= ?
                  AND mp2.fecha <= ?
            )
        ),
        precios_finales AS (
            SELECT 
                mp1.maestro_id, 
                mp1.valor as precio_final,
                mp1.fecha as fecha_final
            FROM maestro_precios mp1
            WHERE mp1.fecha = (
                SELECT MAX(mp2.fecha)
                FROM maestro_precios mp2
                WHERE mp2.maestro_id = mp1.maestro_id
                  AND mp2.fecha >= ?
                  AND mp2.fecha <= ?
            )
        )
        SELECT 
            m.id,
            m.nombre,
            m.unidad,
            pi.precio_inicial,
            pf.precio_final,
            CASE 
                WHEN pi.precio_inicial > 0 THEN
                    ((pf.precio_final - pi.precio_inicial) / pi.precio_inicial * 100.0)
                ELSE 0.0
            END as variacion_percent
        FROM maestro m
        INNER JOIN precios_iniciales pi ON m.id = pi.maestro_id
        INNER JOIN precios_finales pf ON m.id = pf.maestro_id
        WHERE m.tipo = 'P' AND m.activo = 1
        ORDER BY variacion_percent """ + ("DESC" if order_by == "desc" else "ASC")
    
    params = (fecha_desde, fecha_hasta, fecha_desde, fecha_hasta)
    results = execute_query(query, params)
    
    # Format response to match expected structure
    formatted_results = []
    for row in results:
        formatted_results.append({
            'id': row['id'],
            'nombre': row['nombre'],
            'unidad': row['unidad'],
            'precio_inicial': row['precio_inicial'],
            'precio_final': row['precio_final'],
            'variacion_percent': row['variacion_percent']
        })
    
    return jsonify(formatted_results)


@bp.route('/stats/<int:product_id>', methods=['GET'])
def get_product_stats(product_id: int):
    """Get statistics for a product."""
    fecha_desde = request.args.get('fecha_desde', type=str)
    fecha_hasta = request.args.get('fecha_hasta', type=str)
    
    # Convert string dates to date objects if provided
    if fecha_desde:
        fecha_desde = date.fromisoformat(fecha_desde)
    if fecha_hasta:
        fecha_hasta = date.fromisoformat(fecha_hasta)
    
    # Build WHERE clause
    where_clause = "WHERE maestro_id = ?"
    params = [product_id]
    
    if fecha_desde:
        where_clause += " AND fecha >= ?"
        params.append(fecha_desde)
    if fecha_hasta:
        where_clause += " AND fecha <= ?"
        params.append(fecha_hasta)
    
    # Get min, max, and latest price
    query = f"""
        SELECT 
            MIN(valor) as precio_minimo,
            MAX(valor) as precio_maximo,
            MIN(fecha) as fecha_minima,
            MAX(fecha) as fecha_maxima,
            (SELECT valor FROM maestro_precios 
             WHERE maestro_id = ? {f"AND fecha <= ?" if fecha_hasta else ""} 
             ORDER BY fecha DESC LIMIT 1) as precio_actual
        FROM maestro_precios
        {where_clause}
    """
    
    if fecha_hasta:
        params_for_actual = [product_id, fecha_hasta]
    else:
        params_for_actual = [product_id]
    
    # Execute query with all params
    result = execute_query_single(query, tuple(params + params_for_actual))
    
    if not result:
        abort(404, description="Product not found or no data")
    
    # Calculate variation if we have date range
    variacion_periodo = None
    if fecha_desde and fecha_hasta and result.get('precio_minimo') is not None:
        # Get first and last price in range
        first_query = """
            SELECT valor FROM maestro_precios
            WHERE maestro_id = ? AND fecha >= ? AND fecha <= ?
            ORDER BY fecha ASC LIMIT 1
        """
        last_query = """
            SELECT valor FROM maestro_precios
            WHERE maestro_id = ? AND fecha >= ? AND fecha <= ?
            ORDER BY fecha DESC LIMIT 1
        """
        first_result = execute_query_single(first_query, (product_id, fecha_desde, fecha_hasta))
        last_result = execute_query_single(last_query, (product_id, fecha_desde, fecha_hasta))
        
        if first_result and last_result and first_result.get('valor') and last_result.get('valor'):
            first_price = first_result['valor']
            last_price = last_result['valor']
            if first_price > 0:
                variacion_periodo = ((last_price - first_price) / first_price) * 100.0
    
    return jsonify({
        'current_price': result.get('precio_actual'),
        'variation_last_month': variacion_periodo,
        'min_price_period': result.get('precio_minimo'),
        'max_price_period': result.get('precio_maximo')
    })


@bp.route('/products/prices/export', methods=['GET'])
def export_prices_to_excel():
    """Export prices for multiple products to Excel."""
    # Get product_ids from query params (can be multiple with same key)
    product_ids = request.args.getlist('product_ids[]', type=int)
    
    if not product_ids:
        abort(400, description="At least one product_id is required")
    
    fecha_desde = request.args.get('fecha_desde', type=str)
    fecha_hasta = request.args.get('fecha_hasta', type=str)
    
    # Convert string dates to date objects if provided
    if fecha_desde:
        fecha_desde = date.fromisoformat(fecha_desde)
    if fecha_hasta:
        fecha_hasta = date.fromisoformat(fecha_hasta)

    # Build query with IN clause
    placeholders = ",".join("?" * len(product_ids))
    
    if fecha_desde and fecha_hasta:
        query = f"""
            SELECT mp.maestro_id, mp.fecha, mp.valor,
                   m.id, m.nombre, m.tipo, m.unidad, m.categoria, m.fuente, m.periodicidad
            FROM maestro_precios mp
            JOIN maestro m ON mp.maestro_id = m.id
            WHERE mp.maestro_id IN ({placeholders}) 
              AND mp.fecha BETWEEN ? AND ?
            ORDER BY mp.maestro_id, mp.fecha ASC
        """
        params = tuple(product_ids) + (fecha_desde, fecha_hasta)
    elif fecha_desde:
        query = f"""
            SELECT mp.maestro_id, mp.fecha, mp.valor,
                   m.id, m.nombre, m.tipo, m.unidad, m.categoria, m.fuente, m.periodicidad
            FROM maestro_precios mp
            JOIN maestro m ON mp.maestro_id = m.id
            WHERE mp.maestro_id IN ({placeholders}) 
              AND mp.fecha >= ?
            ORDER BY mp.maestro_id, mp.fecha ASC
        """
        params = tuple(product_ids) + (fecha_desde,)
    elif fecha_hasta:
        query = f"""
            SELECT mp.maestro_id, mp.fecha, mp.valor,
                   m.id, m.nombre, m.tipo, m.unidad, m.categoria, m.fuente, m.periodicidad
            FROM maestro_precios mp
            JOIN maestro m ON mp.maestro_id = m.id
            WHERE mp.maestro_id IN ({placeholders}) 
              AND mp.fecha <= ?
            ORDER BY mp.maestro_id, mp.fecha ASC
        """
        params = tuple(product_ids) + (fecha_hasta,)
    else:
        query = f"""
            SELECT mp.maestro_id, mp.fecha, mp.valor,
                   m.id, m.nombre, m.tipo, m.unidad, m.categoria, m.fuente, m.periodicidad
            FROM maestro_precios mp
            JOIN maestro m ON mp.maestro_id = m.id
            WHERE mp.maestro_id IN ({placeholders})
            ORDER BY mp.maestro_id, mp.fecha ASC
        """
        params = tuple(product_ids)

    results = execute_query(query, params)
    
    # Group by product
    products_dict = {}
    for row in results:
        product_id = row['maestro_id']
        if product_id not in products_dict:
            products_dict[product_id] = {
                'product_id': row['id'],
                'product_name': row['nombre'],
                'unit': row['unidad'],
                'data': []
            }
        products_dict[product_id]['data'].append({
            'fecha': row['fecha'],
            'valor': row['valor']
        })
    
    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Precios Corrientes"
    
    # Header style
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Data style
    data_alignment = Alignment(horizontal="right", vertical="center")
    
    # Write headers
    row = 1
    ws['A1'] = 'Producto'
    ws['B1'] = 'Unidad'
    ws['C1'] = 'Fecha'
    ws['D1'] = 'Precio'
    
    # Apply header styles
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    # Write data
    row = 2
    for product_id, product_data in products_dict.items():
        product_name = product_data['product_name']
        unit = product_data['unit'] or ''
        
        for price_data in product_data['data']:
            ws[f'A{row}'] = product_name
            ws[f'B{row}'] = unit
            ws[f'C{row}'] = price_data['fecha']
            ws[f'D{row}'] = price_data['valor']
            
            # Apply data alignment to price column
            ws[f'D{row}'].alignment = data_alignment
            
            row += 1
    
    # Auto-adjust column widths
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    
    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Generate filename
    fecha_str = ""
    if fecha_desde and fecha_hasta:
        fecha_str = f"_{fecha_desde}_{fecha_hasta}"
    elif fecha_desde:
        fecha_str = f"_desde_{fecha_desde}"
    elif fecha_hasta:
        fecha_str = f"_hasta_{fecha_hasta}"
    
    filename = f"precios_corrientes{fecha_str}.xlsx"
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )
