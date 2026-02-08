"""
Exportar IPC de Brasil a Excel
"""
import sys
sys.path.insert(0, 'backend')

from app.database import execute_query, execute_query_single
from openpyxl import Workbook
from datetime import date

print("="*80)
print("EXPORTANDO IPC DE BRASIL A EXCEL")
print("="*80)
print()

# 1. Buscar IPC de Brasil
print("1. Buscando IPC de Brasil...")
query_ipc = """
    SELECT id, nombre, pais, periodicidad
    FROM maestro
    WHERE periodicidad = 'M'
    AND (
        nombre LIKE '%IPC%' OR nombre LIKE '%índice de precios%' OR nombre LIKE '%indice de precios%'
    )
    AND (pais = 'Brasil' OR pais LIKE '%Brasil%')
    AND (activo = 1 OR CAST(activo AS INTEGER) = 1)
    ORDER BY nombre
    LIMIT 1
"""
ipc_info = execute_query_single(query_ipc)

if not ipc_info:
    print("ERROR: No se encontró IPC para Brasil")
    exit(1)

ipc_id = ipc_info['id']
ipc_nombre = ipc_info['nombre']
ipc_pais = ipc_info['pais']

print(f"IPC encontrado:")
print(f"  ID: {ipc_id}")
print(f"  Nombre: {ipc_nombre}")
print(f"  País: {ipc_pais}")

# 2. Obtener todos los datos de IPC de Brasil
print("\n2. Obteniendo datos de IPC...")
query_precios = """
    SELECT fecha, valor
    FROM maestro_precios
    WHERE maestro_id = ?
    ORDER BY fecha ASC
"""
precios = execute_query(query_precios, (ipc_id,))
print(f"Total registros encontrados: {len(precios)}")

if len(precios) == 0:
    print("ERROR: No hay datos de precios para este IPC")
    exit(1)

# 3. Crear Excel
print("\n3. Creando Excel...")
wb = Workbook()
ws = wb.active
ws.title = "IPC Brasil"

# Encabezados
ws['A1'] = 'Fecha'
ws['B1'] = 'Valor'

# Estilo para encabezados
from openpyxl.styles import Font, PatternFill, Alignment
header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=12)

ws['A1'].fill = header_fill
ws['A1'].font = header_font
ws['A1'].alignment = Alignment(horizontal='center')
ws['B1'].fill = header_fill
ws['B1'].font = header_font
ws['B1'].alignment = Alignment(horizontal='center')

# Datos
row = 2
for precio in precios:
    fecha_val = precio.get('fecha')
    valor = precio.get('valor')
    
    # Convertir fecha a string si es necesario
    if isinstance(fecha_val, date):
        fecha_str = fecha_val.strftime('%Y-%m-%d')
    elif isinstance(fecha_val, str):
        fecha_str = fecha_val.split(' ')[0] if ' ' in fecha_val else fecha_val
    else:
        fecha_str = str(fecha_val)
    
    ws[f'A{row}'] = fecha_str
    ws[f'B{row}'] = float(valor) if valor is not None else ''
    row += 1

# Ajustar ancho de columnas
ws.column_dimensions['A'].width = 15
ws.column_dimensions['B'].width = 15

# Guardar
filename = 'ipc_brasil.xlsx'
wb.save(filename)
print(f"\n[OK] Excel creado: {filename}")
print(f"Total filas (incluyendo encabezado): {row}")

print()
print("="*80)
print("FIN")
print("="*80)
